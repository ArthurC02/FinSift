"""
Extract account values from a folder of converted markdown (.md) financial
statements, using a dedicated account-coding dictionary (an .xlsx, e.g.
"Account Coding.xlsx") instead of keyword/text matching. Since every account
line already has a fixed code (e.g. "19999", "A00010", "3110") tied 1:1 to
its account name, a table row is identified in the source document by its
leading code cell matching the dictionary exactly - no fuzzy/substring
matching is needed.

Given a document and a statement type, this pulls every code+value pair
found in that statement in one pass ("whole statement at once"), rather than
one account per run.

Usage:
    python acctfinder.py <folder> <statement> --coding "Account Coding.xlsx" [--period 2024/12/31] [--export csv] [-v]

<statement> is one of: balance_sheet, income_statement, cash_flow
(the fourth sheet in the coding workbook, the equity statement / 權益變動表,
has a transposed layout - codes as column headers rather than row leaders -
and is not supported by this tool; requesting it raises a clear error).

Requires openpyxl to read the coding dictionary (pip install openpyxl).
"""

import argparse
import csv
import re
import sys
from pathlib import Path
from core.text import despace_cjk, _contains_any, _is_toc_like, page_num, strip_footnote, _strip_footnote_suffix
from core.numbers import parse_numeric, nth_value, format_value, annualize, format_pct, format_maybe_pct
from core.industry import INDUSTRY_CODING_FILES, INDUSTRY_CATEGORY_KEYWORDS, detect_industry_category
from core.lookup import find_value_by_label, find_code_value, build_code_index
from core.tables import percent_stride_map, build_raw_lines, _split_dual_column_tables, restrict_section, _is_table_divider, _split_row, _looks_like_code, group_rows_by_code, parse_pipe_tables


def pick_folder():
    """Open a native folder-selection dialog (tkinter, always bundled with
    Python on Windows) and return the chosen path, or None if the user
    cancels. Used as a fallback when <folder> is omitted from the command
    line, so a folder can be picked by clicking instead of typing/pasting
    its full path. Opens in Downloads by default - Windows' native dialog
    otherwise remembers whatever directory some OTHER, unrelated program
    last used it in, which can land somewhere unhelpful (e.g. an old
    project's folder) rather than where these .md conversions actually live."""
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    downloads = Path.home() / "Downloads"
    folder = filedialog.askdirectory(
        title="Select the .md folder",
        initialdir=str(downloads) if downloads.is_dir() else None,
    )
    root.destroy()
    return folder or None

# Windows consoles often default to cp1252, which can't print CJK output.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Statement definitions: coding-workbook sheet name, section markers used to
# isolate that statement's rows within a merged .md document (a single file
# may contain multiple statements back-to-back), and where the code/name
# columns live in that sheet (0-indexed, (original_code_cols, original_name_col,
# revised_code_cols, revised_name_col)).
# ---------------------------------------------------------------------------

STATEMENTS = {
    "balance_sheet": {
        "sheet": "資產負債表",
        "start_markers": ["合併資產負債表", "資產負債表", "Balance Sheet", "BALANCE SHEET"],
        "end_markers": ["合併綜合損益表", "綜合損益表", "Income Statement"],
    },
    "income_statement": {
        "sheet": "綜合損益表",
        "start_markers": ["合併綜合損益表", "綜合損益表", "Income Statement", "Statement of Comprehensive Income"],
        "end_markers": ["合併現金流量表", "現金流量表", "Cash Flow", "CASH FLOW"],
    },
    "cash_flow": {
        "sheet": "現金流量表",
        "start_markers": ["合併現金流量表", "現金流量表", "Cash Flow Statement", "CASH FLOW"],
        "end_markers": ["合併權益變動表", "權益變動表", "Statement of Changes in Equity"],
    },
}

UNSUPPORTED_STATEMENT_MSG = (
    "'equity_statement' (權益變動表) is not supported: its coding sheet is "
    "transposed (codes are column headers, not row leaders), which this "
    "tool's row-per-account matching can't handle."
)

# Industry detection and the coding-workbook table moved to
# core/industry.py - summary mode needs them too. Re-exported above.


def _unmerge_fill(ws):
    """Return a 2D grid of cell values with every merged range's value
    copied into each of its member cells - openpyxl otherwise returns None
    for all but a merge's top-left cell, which would hide most of these
    sheets' header/title text (the industry coding workbooks make heavy
    use of merged header cells spanning several columns)."""
    max_r, max_c = ws.max_row, ws.max_column
    grid = [[ws.cell(row=r, column=c).value for c in range(1, max_c + 1)] for r in range(1, max_r + 1)]
    for mc in ws.merged_cells.ranges:
        val = grid[mc.min_row - 1][mc.min_col - 1]
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                grid[r - 1][c - 1] = val
    return grid


# Header markers distinguishing the "original/pre-revision" code+name block
# from the "revised/post-revision" one (both are present side by side in
# each sheet - a reconciliation between the old and new, IFRS17-era, code
# schemes). Deliberately excludes anything that could ALSO be a real
# account name (e.g. '資產'/'收益' are legitimate level-1 line-item names,
# not just header labels) - an early version of this used those as header
# hints and that caused a real DATA row (code 10000 = '資產') to be misread
# as a header row, pushing the data-start row past it and silently
# dropping that code from the dictionary.
_CODING_ORIG_MARKERS = ["原會計項目及代碼", "修正前會計項目及代碼", "修訂前現金流量表項目及代碼", "修正前"]
_CODING_REV_MARKERS = ["修正後會計項目及代碼", "修訂後現金流量表項目及代碼", "修正後"]
_CODING_HEADER_HINTS = ["代碼", "會計項目", "一級", "二級", "三級", "四級"]
# Columns after these headers are explanatory prose (why a code changed, a
# free-text remark), not code/name data - including them in a block's
# column span lets their long text get picked up as the "name" (via the
# longest-non-code-text heuristic below) instead of the real, much
# shorter, account name.
_CODING_STOP_MARKERS = ["修正說明", "備註"]


def _find_coding_blocks(grid, scan_rows=10):
    """Scan the first `scan_rows` rows for the original-block and revised-
    block column spans (see marker lists above). Returns (orig_span,
    rev_span) as 0-indexed (start, end) inclusive tuples, or None for a
    block absent from this sheet, plus the 0-indexed row where real data
    starts (the row after the last header-hint row found)."""
    n_cols = max(len(r) for r in grid[:scan_rows])
    orig_start = rev_start = stop_col = None
    last_header_row = 0
    for r in range(min(scan_rows, len(grid))):
        row_has_hint = False
        for c in range(n_cols):
            val = grid[r][c]
            if not isinstance(val, str):
                continue
            if any(m in val for m in _CODING_ORIG_MARKERS) and orig_start is None:
                orig_start = c
                row_has_hint = True
            if any(m in val for m in _CODING_REV_MARKERS) and rev_start is None:
                rev_start = c
                row_has_hint = True
            if any(m in val for m in _CODING_STOP_MARKERS) and stop_col is None:
                stop_col = c
                row_has_hint = True
            if val.strip() in _CODING_HEADER_HINTS:
                row_has_hint = True
        if row_has_hint:
            last_header_row = r
    limit = (stop_col - 1) if stop_col is not None else (n_cols - 1)
    orig_end = (rev_start - 1) if (orig_start is not None and rev_start is not None and rev_start > orig_start) else limit
    orig_block = (orig_start, orig_end) if orig_start is not None else None
    rev_block = (rev_start, limit) if rev_start is not None else None
    return orig_block, rev_block, last_header_row + 1


def _extract_coding_block(grid, block, data_start_row):
    """For a (start, end) column span, build {code: name} by scanning each
    data row's cells in that span: the code-shaped cell is the code, and
    the longest non-code text cell is the name (not a fixed column index -
    the name column shifts between rows in at least one observed sheet)."""
    if block is None:
        return {}
    start, end = block
    out = {}
    for row in grid[data_start_row:]:
        cells = row[start:end + 1]
        code = name = None
        for cell in cells:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            if code is None and _looks_like_code(s):
                code = s
            elif not _looks_like_code(s) and (name is None or len(s) > len(name)):
                name = s
        if code and name:
            out[code] = name
    return out


def load_code_dictionary(coding_path, statement):
    """Return {code_str: name} for the given statement, read from an
    industry coding workbook (see INDUSTRY_CODING_FILES). Both the
    original and revised code/name blocks are read; revised entries win
    on conflict since 修正後 is the corrected version. Blank code cells
    are skipped."""
    if openpyxl is None:
        raise RuntimeError("Reading the coding dictionary requires 'openpyxl'. Install with: pip install openpyxl")
    if statement == "equity_statement":
        raise ValueError(UNSUPPORTED_STATEMENT_MSG)
    if statement not in STATEMENTS:
        raise ValueError(f"Unknown statement '{statement}'. Known: {list(STATEMENTS)}")

    spec = STATEMENTS[statement]
    wb = openpyxl.load_workbook(coding_path, data_only=True)
    if spec["sheet"] not in wb.sheetnames:
        raise ValueError(f"Sheet '{spec['sheet']}' not found in {coding_path}. Sheets: {wb.sheetnames}")
    ws = wb[spec["sheet"]]

    grid = _unmerge_fill(ws)
    orig_block, rev_block, data_start_row = _find_coding_blocks(grid)
    orig_dict = _extract_coding_block(grid, orig_block, data_start_row)
    rev_dict = _extract_coding_block(grid, rev_block, data_start_row)
    codes = dict(orig_dict)
    codes.update(rev_dict)
    return codes


def resolve_coding_path(industry=None, folder=None, explicit_path=None):
    """Resolve which coding workbook to load: an explicit path always wins;
    otherwise resolve `industry` (a INDUSTRY_CODING_FILES key) directly, or
    auto-detect it from `folder` if industry is None. Raises ValueError if
    neither an explicit path nor a resolvable industry is available."""
    if explicit_path:
        return explicit_path
    if industry is None and folder is not None:
        industry = detect_industry_category(folder)
    if industry not in INDUSTRY_CODING_FILES:
        raise ValueError(
            f"Couldn't determine an industry coding file (detected: {industry!r}). "
            f"Pass --coding explicitly, or --industry one of {list(INDUSTRY_CODING_FILES)}."
        )
    return INDUSTRY_CODING_FILES[industry]


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def extract_statement(doc_path, statement, code_dict, period=1, verbose=False):
    """Scan doc_path for the given statement's section, match every row
    whose leading cell is a known code, and pull that row's `period`-th
    (1-indexed, most-recent-first) value. Returns a list of dicts:
      {code, name, label_in_doc, value, page_num, confidence}
    (possibly empty if nothing in this file matches).

    If the statement's section marker (e.g. "資產負債表") isn't found in
    this file, the WHOLE file is scanned unrestricted rather than skipped -
    some conversions split one statement per file without repeating its
    title on the data page itself (confirmed against a real 北富銀 filing,
    where the title only appears in the table of contents), so requiring
    the marker would silently produce zero results for an entire folder of
    otherwise-correct data. Matching is by exact code equality regardless,
    so scanning unrestricted doesn't risk false positives the way a looser
    text search would.
    """
    spec = STATEMENTS[statement]
    lines = build_raw_lines(doc_path)
    restricted = restrict_section(lines, spec["start_markers"], spec["end_markers"])
    if restricted is not None:
        lines = restricted

    entries = group_rows_by_code(lines, code_dict)
    results = []
    seen_codes = set()
    for code, page_num, cells, stride in entries:
        if code in seen_codes:
            continue  # keep first match only
        seen_codes.add(code)
        value = nth_value(cells, period, stride)
        results.append({
            "code": code,
            "name": code_dict[code],
            "label_in_doc": cells[1] if len(cells) > 1 else "",
            "value": value,
            "page_num": page_num,
            "confidence": "high" if value is not None else "low",
        })
        if verbose and value is None:
            print(f"  [{code}] matched but no period-{period} value found in row: {cells}")
    return results


_ROC_DATE_RE = re.compile(r"(\d{2,3})\s*年\s*(\d{1,2})\s*月")


def derive_quarter_num(doc_path):
    """Parse the first ROC-calendar 'NNN年N月' date mentioned in the
    document (its title/header prose gives the current period's month) and
    return which quarter that month falls in (1-4), or None if no such date
    is found."""
    text = doc_path.read_text(encoding="utf-8", errors="ignore")
    m = _ROC_DATE_RE.search(text)
    if not m:
        return None
    month = int(m.group(2))
    return (month - 1) // 3 + 1


# ---------------------------------------------------------------------------
# Reported profitability table (獲利能力): ROA/ROE as directly disclosed by
# the filer, rather than computed. This is the authoritative source when
# present - a dedicated section listing 資產報酬率 (ROA) and 淨值報酬率 (ROE),
# each split into 稅前/稅後 (pretax/posttax), plus 純益率 (profit margin), for
# the consolidated group (合併) and every subsidiary, for the current quarter
# and the same quarter last year. Only 稅後 (posttax) figures are surfaced in
# the output (pretax is parsed internally where needed to keep column
# position correct, but isn't reported) - both the as-disclosed (cumulative
# year-to-date) figure and an annualized version (x 4/quarter number) are
# shown, since the disclosed figure is NOT itself annualized (e.g. a Q1
# filing's ROA reflects only ~1 quarter of return).
#
# Two layouts are supported, both seen in real filings:
#
# 1. Row = entity, column = metric (verified against real Cathay .md data,
#    where the header is plain prose above the table, not a real table
#    header row):
#      | 合併獲利能力 | 0.27 | 0.22 | 5.13 | 4.23 | 43.64 |
#      | 本 公 司 | 3.25 | 3.29 | 4.20 | 4.25 | 98.29 |
#    Column order is fixed by the regulatory disclosure format: ROA稅前,
#    ROA稅後, ROE稅前, ROE稅後, 純益率. Period (current quarter vs. same
#    quarter last year) is a block-level distinction (one table per period,
#    introduced by a ROC date-RANGE prose line). The 合併獲利能力
#    (consolidated) row is always cleanly pipe-delimited even where
#    subsidiary rows below it get corrupted by a stray parenthesis/pipe
#    artifact, so entity rows are reconstructed via continuation-folding
#    (like account-code rows) and values are picked by scanning tokens, not
#    trusting fixed cell counts.
#
# 2. Row = metric, column = period (seen in an E.Sun/玉山金控 filing, one
#    small table per entity, each introduced by a numbered heading line like
#    "1. 玉山金控及子公司"):
#      項目          114年12月31日  113年12月31日
#      資產報酬率  稅前   0.95          0.84
#                  稅後   0.80          0.68
#      淨值報酬率  稅前   15.54         13.17
#                  稅後   13.05         10.68
#      純益率      37.48         34.34
#    Entity is a block-level distinction (one table per entity); period is
#    the column distinction, identified from single ROC dates in the header.
#    NOTE: this path is built from that filing's raw PDF text (via pypdf),
#    not yet verified against the actual .md conversion for this layout -
#    treat as provisional pending validation against real converted output.
#    It assumes (per user confirmation) that a real header row is present
#    inside the table for this layout, unlike layout 1 above.
#
# Orientation is detected per-table from header content: a header containing
# ROC dates -> layout 2 (period columns); a header/first-row containing
# ROA/ROE/稅前/稅後-style text -> layout 1 (metric columns, entity rows).
# ---------------------------------------------------------------------------


_PROFITABILITY_SECTION_RE = re.compile(r"獲利能力")

# The company-keyword branch is anchored: an ordinary account name like
# 存放銀行同業 contains 銀行 but continues past it, and used to be read as an
# entity row - swallowing the rows beneath it as continuations. The first
# three alternatives stay PREFIX matches on purpose, so labels like
# 本公司及子公司 keep working; anchoring those too would trade this false
# positive for a false negative, which corrupts the primary path instead.
_ENTITY_ROW_RE = re.compile(r"^(合併|本\s*公\s*司|國\s*泰|.+(?:銀行|人壽|產險|證券|保險)(?:股份有限公司|公司)?$)")

_PERIOD_RANGE_RE = re.compile(
    r"(\d{2,3})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日\s*至\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"
)

# Metric tokens are small decimals (not comma-grouped), optionally negative
# in parens, or a bare "-" placeholder for "not available" - a bare "-"
# still consumes a position so the fixed ROA稅前/稅後/ROE稅前/稅後/純益率
# column order is preserved even when one metric is missing.
# The placeholder alternative covers the full-width dashes too: a "—" used to
# match nothing, so it held no position and every metric after it in the row
# was read one column early.
_METRIC_TOKEN_RE = re.compile(r"\(?\s*-?\d+(?:\.\d+)?\s*\)?|[-\u2014\u2013]")

_METRIC_NAMES = ["roa_pretax", "roa_posttax", "roe_pretax", "roe_posttax", "profit_margin"]

# Single ROC date, e.g. "114年12月31日" - used both to find period-label
# months (for annualization) and, in layout 2, to identify a table's period
# header columns.
_SINGLE_DATE_RE = re.compile(r"(\d{2,3})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日")

# Numbered heading line introducing one entity's table in layout 2, e.g.
# "1. 玉山金控及子公司" (after despacing).
_ENTITY_HEADING_RE = re.compile(r"^\s*\d+[.．]\s*(.+?)\s*$")

_ROW_METRIC_PATTERNS = {
    "roa_posttax": (re.compile(r"資產報酬率"), re.compile(r"稅\s*後")),
    "roa_pretax": (re.compile(r"資產報酬率"), re.compile(r"稅\s*前")),
    "roe_posttax": (re.compile(r"淨值報酬率"), re.compile(r"稅\s*後")),
    "roe_pretax": (re.compile(r"淨值報酬率"), re.compile(r"稅\s*前")),
}
_PROFIT_MARGIN_ROW_RE = re.compile(r"純益率")


def quarter_num_from_period_label(label):
    """Derive which quarter (1-4) a period label refers to. Handles both a
    ROC date-range label ('115年1月1日至3月31日' - uses the END month) and a
    single ROC date label ('114年12月31日' - uses that month)."""
    if not label:
        return None
    m = _PERIOD_RANGE_RE.search(label)
    if m:
        return (int(m.group(4)) - 1) // 3 + 1
    m2 = _SINGLE_DATE_RE.search(label)
    if m2:
        return (int(m2.group(2)) - 1) // 3 + 1
    return None


def find_profitability_files(folder):
    """Return sorted list of .md files under `folder` whose text mentions a
    獲利能力 section at all (a loose first pass; files without an actual
    data table are filtered out later when no entity rows are found)."""
    return [p for p in sorted(Path(folder).rglob("*.md"))
            if _PROFITABILITY_SECTION_RE.search(p.read_text(encoding="utf-8", errors="ignore"))]


def is_metric_column_layout(lines):
    """Layout 1 check: does the first pipe-table row look like an entity
    name (row=entity, column=metric - the verified Cathay-style layout)?"""
    for _page_num, line in lines:
        if "|" not in line or _is_table_divider(line):
            continue
        cells = _split_row(line)
        if not cells:
            continue
        first = strip_footnote(cells[0])
        return bool(_ENTITY_ROW_RE.match(first))
    return False


def parse_single_date(cell):
    m = _SINGLE_DATE_RE.search(cell)
    if not m:
        return None
    return (int(m.group(1)), int(m.group(2)), int(m.group(3)))


_DOT_DATE_RE = re.compile(r"^(\d{2,3})\.(\d{1,2})\.(\d{1,2})$")
# A bare fiscal year, no month/day at all ('114年度' - confirmed in real
# 國泰/北富銀 filings) - always the full year, so keyed as its year-end
# date (year, 12, 31), the same (year, month, day) shape every other
# format here resolves to.
_FISCAL_YEAR_RE = re.compile(r"^(\d{2,3})\s*年度$")


def parse_period_header_date(cell):
    """Parse a profitability-table column header into a (year, month, day)
    sort key, trying every header date format confirmed in a real filing:
    a single ROC year-month-day ('114年12月31日' - parse_single_date), a
    dot-separated ROC date ('114.12.31' - a real 中信 filing), a period
    RANGE ('115年1月1日至3月31日' - a real 國泰 filing, keyed off its END
    date, the same convention quarter_num_from_period_label uses), or a
    bare fiscal year ('114年度' - a real 國泰/北富銀 filing). None if the
    cell matches none of these."""
    cell = cell.strip()
    # Checked BEFORE parse_single_date: _SINGLE_DATE_RE is unanchored
    # (.search(), not .match()), so on a RANGE string like '114年1月1日至
    # 3月31日' it would spuriously match just the embedded START date
    # ('114年1月1日') and silently mislabel the period by its first day
    # instead of its correct quarter-end - confirmed on a real 國泰/北富銀
    # Q1 filing (value still came out right, since both the current and
    # prior-year columns got the same wrong-but-consistent treatment, but
    # the displayed period label was wrong).
    m = _PERIOD_RANGE_RE.search(cell)
    if m:
        return (int(m.group(1)), int(m.group(4)), int(m.group(5)))
    key = parse_single_date(cell)
    if key:
        return key
    m = _DOT_DATE_RE.match(cell)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = _FISCAL_YEAR_RE.match(cell)
    if m:
        return (int(m.group(1)), 12, 31)
    return None


# A numbered heading only names an ENTITY if it reads like a company name.
# '1. 玉山金控及子公司' does; '1. 前言' and '2. 重要會計政策之說明' do not, yet
# _ENTITY_HEADING_RE matches any numbered line at all - so a 前言 on page one
# used to make every layout-3 table in the file look like layout 2's, and
# layout 3 skipped the lot.
# Duplicates callfinder._ENTITY_NAME_RE's vocabulary. Both belong in core/,
# but moving them is a refactor and this is a bug-fix commit - see the note
# added to TEST_DESIGN §7.
_ENTITY_HEADING_NAME_RE = re.compile(
    r"銀行|金控|控股|人壽|產險|保險|證券|投信|投顧|公司|Bank|FHC|Holdings|Financial|Life|Securities|Insurance")


def _has_entity_heading_before(lines, line_idx):
    """True if some earlier line is a numbered heading that names a company -
    i.e. a table layout 2 has already claimed, which layout 3 must not
    double-count.

    Deliberately narrows only THIS check, not layout 2's own attribution: if
    the two disagree about a heading, the same table is extracted twice with
    the same figures, once under the heading's text and once with entity=None.
    _select_profitability_entry prefers the entity=None copy (always in
    scope), so the worst case is a redundant entry, never a lost one.
    """
    for k in range(line_idx - 1, -1, -1):
        m = _ENTITY_HEADING_RE.match(lines[k][1])
        if m and _ENTITY_HEADING_NAME_RE.search(m.group(1)):
            return True
    return False


def extract_single_entity_profitability_tables(doc_path, verbose=False):
    """Layout 3 extractor: a flat single-entity 獲利能力 table - the shape
    actually used by every real INDIVIDUAL (個體) bank filing's own
    footnote disclosure seen so far (confirmed in real 中信/國泰/玉山
    filings), unlike layout 1 (entity-name row labels) or layout 2 (a
    numbered '1. some entity' heading before the table) which both assume
    a multi-entity comparison table. Here there is exactly one entity - the
    filing's own bank - so every row is attributed to it (entity: None;
    the caller already knows which bank the folder belongs to). Skips any
    table layout 2 already claimed (a numbered entity heading precedes it),
    so the same table is never double-counted under both layouts. Returns
    a list of {entity: None, period_label, quarter_num, roa_posttax,
    roe_posttax, profit_margin, source_file}."""
    lines = build_raw_lines(doc_path)
    results = []
    for table in parse_pipe_tables(lines):
        if _has_entity_heading_before(lines, table["line_idx"]):
            continue
        header = table["header"]
        period_cols = [(idx, parse_period_header_date(cell)) for idx, cell in enumerate(header)]
        period_cols = [(idx, key) for idx, key in period_cols if key is not None]
        if not period_cols:
            continue  # not a profitability (or any date-columned) table
        if not any(classify_metric_row(row) for row in table["rows"]):
            continue  # has date columns, but no ROA/ROE/profit-margin row - unrelated table
        period_cols.sort(key=lambda t: t[1], reverse=True)

        per_period = {idx: {} for idx, _key in period_cols}
        for row in table["rows"]:
            metric = classify_metric_row(row)
            if metric is None:
                continue
            for idx, _key in period_cols:
                if idx < len(row):
                    per_period[idx][metric] = parse_numeric(row[idx])

        for idx, (year, month, day) in period_cols:
            metrics = per_period[idx]
            results.append({
                "entity": None,
                "period_label": f"{year}年{month}月{day}日",
                "quarter_num": (month - 1) // 3 + 1,
                "roa_posttax": metrics.get("roa_posttax"),
                "roe_posttax": metrics.get("roe_posttax"),
                "profit_margin": metrics.get("profit_margin"),
                "source_file": doc_path.name,
            })
        if verbose and period_cols:
            print(f"[{doc_path.name}] layout-3 (single-entity) profitability table: {len(period_cols)} period(s)")
    return results


def classify_metric_row(cells):
    """Return one of 'roa_posttax', 'roa_pretax', 'roe_posttax',
    'roe_pretax', 'profit_margin', or None, from a layout-2 row's leading
    label cell(s) (metric name and 稅前/稅後 may be split across 2 cells)."""
    label = " ".join(cells[:2])
    if _PROFIT_MARGIN_ROW_RE.search(label):
        return "profit_margin"
    for name, (metric_re, tax_re) in _ROW_METRIC_PATTERNS.items():
        if metric_re.search(label) and tax_re.search(label):
            return name
    return None


def extract_transposed_entity_tables(doc_path, verbose=False):
    """Layout 2 extractor: metrics as rows, periods as columns, one table
    per entity, entity identified by the nearest preceding numbered heading
    line (e.g. '1. 玉山金控及子公司'). See the module-level comment above for
    the verification caveat on this layout. Returns a list of
    {entity, period_label, quarter_num, roa_posttax, roe_posttax,
    profit_margin, source_file}."""
    lines = build_raw_lines(doc_path)
    results = []
    for table in parse_pipe_tables(lines):
        header = table["header"]
        # parse_period_header_date, not parse_single_date: the latter is
        # unanchored, so on a RANGE header ("115年1月1日至3月31日") it matched
        # the embedded START date and the period was labelled by its first
        # day instead of its quarter end. Layout 3 already used the
        # range-aware parser; this extractor never got the same fix.
        period_cols = [(idx, parse_period_header_date(cell)) for idx, cell in enumerate(header)]
        period_cols = [(idx, key) for idx, key in period_cols if key is not None]
        if not period_cols:
            continue  # not a period-columns (layout 2) table
        period_cols.sort(key=lambda t: t[1], reverse=True)  # most recent period first

        entity = None
        for k in range(table["line_idx"] - 1, -1, -1):
            m = _ENTITY_HEADING_RE.match(lines[k][1])
            if m:
                entity = m.group(1).strip()
                break
        if entity is None:
            continue  # can't attribute this table to an entity; skip rather than guess

        per_period = {idx: {} for idx, _key in period_cols}
        for row in table["rows"]:
            metric = classify_metric_row(row)
            if metric is None:
                continue
            for idx, _key in period_cols:
                if idx < len(row):
                    per_period[idx][metric] = parse_numeric(row[idx])

        for idx, (year, month, day) in period_cols:
            metrics = per_period[idx]
            period_label = f"{year}年{month}月{day}日"
            results.append({
                "entity": entity,
                "period_label": period_label,
                "quarter_num": (month - 1) // 3 + 1,
                "roa_posttax": metrics.get("roa_posttax"),
                "roe_posttax": metrics.get("roe_posttax"),
                "profit_margin": metrics.get("profit_margin"),
                "source_file": doc_path.name,
            })
        if verbose:
            print(f"[{doc_path.name}] transposed profitability table for entity '{entity}': "
                  f"{len(period_cols)} period(s)")
    return results


def group_rows_by_entity(lines):
    """Same continuation-folding idea as group_rows_by_code, but keyed on an
    entity label (footnote-stripped) recognized by _ENTITY_ROW_RE instead of
    a coding-dictionary code, since 獲利能力 entity rows aren't in the coding
    workbook at all."""
    entries = []
    current = None  # (entity, cells)
    for _page_num, line in lines:
        if "|" not in line or _is_table_divider(line):
            continue
        cells = _split_row(line)
        if not cells:
            continue
        first = strip_footnote(cells[0])
        if _ENTITY_ROW_RE.match(first):
            if current is not None:
                entries.append(current)
            current = (first, list(cells))
        elif current is not None:
            current[1].extend(cells)
    if current is not None:
        entries.append(current)
    return entries


def extract_metrics(cells):
    """Return {roa_pretax, roa_posttax, roe_pretax, roe_posttax,
    profit_margin} parsed positionally from an entity row's cells (skipping
    the leading label cell), None for any metric missing or unavailable."""
    text = " ".join(cells[1:])
    tokens = _METRIC_TOKEN_RE.findall(text)
    result = {}
    for i, name in enumerate(_METRIC_NAMES):
        result[name] = parse_numeric(tokens[i]) if i < len(tokens) else None
    return result


def find_profitability_entries(folder, verbose=False):
    """Scan `folder` for 獲利能力 disclosure tables in either supported
    layout. Returns a flat list of {entity, period_label, quarter_num,
    roa_posttax, roe_posttax, profit_margin, source_file}, one entry per
    (entity, period) found. Empty list if no file has either layout's data
    table (a bare mention of 獲利能力 in passing doesn't count)."""
    results = []
    for doc_path in find_profitability_files(folder):
        lines = build_raw_lines(doc_path)

        # Layout 1: row=entity/column=metric, one table per period block
        # (block boundaries are ROC date-RANGE prose lines).
        boundaries = [i for i, (_pn, line) in enumerate(lines) if _PERIOD_RANGE_RE.search(line)]
        for bi, start in enumerate(boundaries):
            end = boundaries[bi + 1] if bi + 1 < len(boundaries) else len(lines)
            block_lines = lines[start:end]
            if not is_metric_column_layout(block_lines):
                continue
            m = _PERIOD_RANGE_RE.search(block_lines[0][1])
            period_label = m.group(0) if m else None
            quarter_num = quarter_num_from_period_label(period_label)
            entries = group_rows_by_entity(block_lines)
            for entity, cells in entries:
                metrics = extract_metrics(cells)
                results.append({
                    "entity": entity, "period_label": period_label, "quarter_num": quarter_num,
                    "roa_posttax": metrics["roa_posttax"], "roe_posttax": metrics["roe_posttax"],
                    "profit_margin": metrics["profit_margin"], "source_file": doc_path.name,
                })
            if entries and verbose:
                print(f"[{doc_path.name}] layout-1 profitability block '{period_label}': {len(entries)} entit(y/ies)")

        # Layout 2: row=metric/column=period, one table per entity.
        results.extend(extract_transposed_entity_tables(doc_path, verbose=verbose))

        # Layout 3: row=metric/column=period, single entity (no numbered
        # heading) - the common shape in an individual (個體) filing.
        results.extend(extract_single_entity_profitability_tables(doc_path, verbose=verbose))
    return results


def compute_ratios(folder, bank, coding_path=None, verbose=False):
    """Compute ROA(稅後年化) and ROE(稅後年化) from the same account codes
    and per-bank override/label-fallback logic SUMMARY_LAYOUT already uses
    (10000=資產, 30000=權益, 64000/63000=稅後淨利 per bank) via
    find_code_value. It used to go through a section-marker-restricted
    lookup (find_statement_rows, since deleted as dead code), which turned
    out to be unreliable in practice (a real 國泰 filing's income-statement
    page never repeats its own '...綜合損益表' section title, so the marker
    search silently found nothing even though the code rows were right
    there) and, separately, used codes (19999/39999/69000) left over from
    before this project's industry-coding-dictionary rewrite - confirmed
    absent from the current 金控業/金融業/保險業 workbooks entirely (that
    mapping was RATIO_CODES, also now deleted). Raises RuntimeError with a
    clear message if any required code/quarter can't be found.
    coding_path is unused now (kept as a parameter for backward
    compatibility with existing callers) since find_code_value doesn't
    need a coding dictionary - it matches raw document codes directly."""
    overrides = SUMMARY_CODE_OVERRIDES.get(bank, {})
    net_income_code = overrides.get("64000", "64000")

    def get(code, period, label, label_fallback=None):
        found = find_code_value(folder, code, period=period, verbose=verbose, label_fallback=label_fallback)
        if found is None:
            raise RuntimeError(f"Code {code} ({label}) not found in any file.")
        _label, value, source_file = found
        if value is None:
            raise RuntimeError(f"Code {code} ({label}) was found but had no period-{period} value.")
        return value, source_file

    net_income, inc_source = get(net_income_code, 1, "本期稅後淨利（淨損）",
                                  SUMMARY_LABEL_FALLBACKS.get("64000"))
    assets_cur, bs_source = get("10000", 1, "資產總計, current quarter", SUMMARY_LABEL_FALLBACKS.get("10000"))
    assets_prev, _ = get("10000", 2, "資產總計, last quarter", SUMMARY_LABEL_FALLBACKS.get("10000"))
    equity_cur, _ = get("30000", 1, "權益總計, current quarter", SUMMARY_LABEL_FALLBACKS.get("30000"))
    equity_prev, _ = get("30000", 2, "權益總計, last quarter", SUMMARY_LABEL_FALLBACKS.get("30000"))

    bs_path = next(Path(folder).rglob(bs_source), None)
    quarter_num = derive_quarter_num(bs_path) if bs_path else None
    if quarter_num is None:
        raise RuntimeError(f"Couldn't determine the current quarter number from {bs_source}.")

    # net_income IS a year-to-date cumulative figure, not a standalone
    # current-quarter one - confirmed directly from real filings' own
    # income-statement column headers: a Q4/annual filing's column reads
    # "114年度金額" (FY114 amount - the FULL YEAR, e.g. a real 國泰 filing),
    # while a Q2 filing's column reads "一月一日至六月三十日" (Jan 1 - Jun
    # 30, i.e. H1 CUMULATIVE, not Q2 alone - a real 中信 filing). An earlier
    # version of this function assumed the opposite (single-quarter, not
    # cumulative) and annualized with a flat x4 regardless of quarter - for
    # a Q4 filing (already the full year) that quadruples an already-annual
    # figure, which is exactly the ~4x-too-high result three different
    # banks' Q4 crosschecks showed before this fix. Dividing by quarter_num
    # first correctly reduces to a no-op at quarter_num=4 (already annual)
    # and projects a partial-year cumulative figure to a full year otherwise.
    # A zero average balance is a data problem of the same class as a missing
    # code, so it has to surface the same way. Letting the division raise
    # ZeroDivisionError instead broke this function's documented contract
    # ("Raises RuntimeError ..."), and since ZeroDivisionError is NOT a
    # subclass of RuntimeError, collect_roa_roe's `except RuntimeError` let it
    # through and took down the whole run rather than degrading to
    # "cross-check unavailable" the way every other lookup failure does.
    avg_assets = (assets_cur + assets_prev) / 2
    avg_equity = (equity_cur + equity_prev) / 2
    for label, denom in (("資產總計 (10000)", avg_assets), ("權益總計 (30000)", avg_equity)):
        if denom == 0:
            raise RuntimeError(f"Average {label} across the two periods is 0 - "
                               f"can't compute a ratio from it.")

    roa = net_income / avg_assets / quarter_num * 4
    roe = net_income / avg_equity / quarter_num * 4

    if verbose:
        print(f"Balance sheet: {bs_source} | Income statement: {inc_source} | quarter: Q{quarter_num}")
        print(f"  net_income ({net_income_code}) = {format_value(net_income)}")
        print(f"  total_assets (10000)  = {format_value(assets_cur)} (current), {format_value(assets_prev)} (last quarter)")
        print(f"  total_equity (30000)  = {format_value(equity_cur)} (current), {format_value(equity_prev)} (last quarter)")

    return {"roa": roa, "roe": roe, "quarter_num": quarter_num}


def _select_profitability_entry(entries, bank):
    """Pick the most-recent, best-scoped entry from find_profitability_entries'
    output. entity=None (layout 3 - a filing's single entity) always counts
    as in-scope, since there's nothing else it could be; otherwise prefer an
    entry whose entity text names `bank` (via BANK_NAME_ALIASES), falling
    back to any entity if none matches. Among the scoped candidates, prefers
    one that actually has a value, then the most recent period. None if
    `entries` is empty."""
    if not entries:
        return None
    aliases = BANK_NAME_ALIASES.get(bank, []) if bank else []
    scoped = [e for e in entries
              if e.get("entity") is None or (aliases and any(a in e["entity"] for a in aliases))]
    pool = scoped or entries
    with_value = [e for e in pool if e["roa_posttax"] is not None or e["roe_posttax"] is not None]
    pool = with_value or pool

    def period_key(e):
        m = parse_single_date(e["period_label"]) if e["period_label"] else None
        return m or (0, 0, 0)

    return max(pool, key=period_key)


# Two as-disclosed ROA/ROE readings for the SAME quarter shouldn't diverge
# by more than this factor if they're measuring the same thing the same
# way; a bigger gap flags a likely convention mismatch (e.g. one figure is
# a raw cumulative-YTD number, the other already annualized) worth a human
# double-check, rather than silently picking one.
_ROA_ROE_CROSSCHECK_DIVERGENCE_FACTOR = 2.0

# Wide, deliberately generous plausibility bounds - catch a grossly
# implausible value (e.g. a decimal/percent parsing bug producing ~92%
# instead of ~0.92%), not legitimate variation. Grounded in real observed
# data across all 4 currently-supported banks this session: ROA(稅後)
# ranged 0.24%-1.12%, ROE(稅後) ranged 3.33%-15.02%. Allows negative
# values too, since a quarterly loss is a real result, not a bug - this is
# a sanity check on magnitude, not a claim about which sign is expected.
_ROA_PLAUSIBLE_MIN, _ROA_PLAUSIBLE_MAX = -5.0, 5.0     # percent
_ROE_PLAUSIBLE_MIN, _ROE_PLAUSIBLE_MAX = -50.0, 50.0   # percent


def collect_roa_roe(folder, bank, coding=None, concall_roa=None, concall_roe=None, verbose=False):
    """ROA/ROE for the curated fin_report summary, in priority order:
      1. the filing's own reported 獲利能力 disclosure table (any of the 3
         layouts - see find_profitability_entries), used AS DISCLOSED, with
         NO further annualizing applied to it. An earlier version of this
         project always scaled the disclosed figure by x4/quarter_num,
         assuming Taiwanese banks uniformly report a not-yet-annualized
         cumulative rate here. Cross-checking real filings for 3 different
         banks across 2 periods each disproved that: 中信 and 玉山's Q1
         disclosed rate was roughly the SAME magnitude as their own Q4
         rate (already annualized), while 國泰's Q1 rate was roughly 1/4 of
         its Q4 rate (genuinely not yet annualized) - i.e. the convention
         is NOT consistent across banks, so blindly scaling is wrong for
         at least some of them. Showing the as-disclosed number, with an
         independent cross-check alongside it (see below), is the only
         choice that doesn't risk silently fabricating a wrong figure.
      2. concall_roa/concall_roe - an earnings-call deck's own reported
         figure, supplied by the caller (runfinder.py) rather than looked
         up here, since this module can't import callfinder.py's term
         matching without a circular dependency.
      3. this fin folder's own manual formula (compute_ratios) as a last
         resort, clearly labeled as an approximation.
    Regardless of which source wins, the manual formula is ALSO computed
    (when derivable) and returned as a cross-check value - never used to
    silently override a disclosed/concall figure - with a note when it
    diverges from the primary value by more than
    _ROA_ROE_CROSSCHECK_DIVERGENCE_FACTOR, since that gap itself is useful
    signal (see above) rather than something to hide.
    Returns {"roa": row_or_None, "roe": row_or_None}, each row a dict with
    term, value, matched_label, source_file, crosscheck_value, note."""
    entries = find_profitability_entries(folder, verbose=verbose)
    entry = _select_profitability_entry(entries, bank)

    manual = None
    try:
        manual = compute_ratios(folder, bank, coding, verbose=verbose)
    except RuntimeError as e:
        if verbose:
            print(f"ROA/ROE manual-formula cross-check unavailable: {e}")

    def build(term, metric_key, concall_value, manual_key):
        crosscheck = manual[manual_key] * 100 if manual else None  # manual formula is a plain ratio, not %
        if entry is not None and entry[metric_key] is not None:
            value, source, label, source_file = (
                entry[metric_key], "fin_report disclosure",
                f"{term}(年) 稅後 @ {entry['period_label']}", entry["source_file"],
            )
        elif concall_value is not None:
            value, source, label, source_file = concall_value, "concall", term, None
        elif crosscheck is not None:
            value, source, label, source_file = crosscheck, "fin_report manual formula (approximated)", term, None
            crosscheck = None  # it IS the value here, not a separate cross-check of itself
        else:
            return None

        notes = []
        if crosscheck is not None and value:
            # Compare MAGNITUDES, and treat opposite signs as divergent
            # outright. The numerator used to be max(value, crosscheck) with
            # no abs(), so a loss quarter - a perfectly ordinary input, the
            # plausible range runs down to -5%/-50% - could produce a ratio of
            # 1.0 or even a negative one and silently pass: max(-3.0, 1.0) is
            # 1.0, max(-3.0, -1.0) is -1.0. Opposite signs mean one source
            # says profit and the other says loss, which is the largest
            # disagreement there is and was likewise never reported.
            magnitude_ratio = max(abs(value), abs(crosscheck)) / min(abs(value), abs(crosscheck) or 1e-9)
            if (value > 0) != (crosscheck > 0) or magnitude_ratio > _ROA_ROE_CROSSCHECK_DIVERGENCE_FACTOR:
                # Don't assert WHY they diverge - the manual formula's own
                # single-quarter-net-income x4 assumption is itself unverified
                # for a Q4/annual filing (where "this period's" net income
                # code may cover the full year, not just Q4 alone), so a gap
                # here isn't reliable evidence the disclosed figure is wrong.
                notes.append(f"cross-check diverges: manual formula gives {crosscheck:.2f}% vs {value:.2f}% "
                             f"as-disclosed - could be a real discrepancy, or just this manual formula's own "
                             f"assumptions not holding for this filing; treat as a prompt to look closer, not "
                             f"as evidence either number is wrong")

        # Plausibility bound - a SEPARATE signal from the cross-check
        # divergence above (that flags disagreement BETWEEN two
        # measurements of the same thing; this flags a value that's
        # implausible in absolute terms regardless of source or agreement).
        # Deliberately wide (see the bounds' own comment), so this should
        # almost never fire on real data - both notes are kept if both fire,
        # rather than one overwriting the other.
        lo, hi = (_ROA_PLAUSIBLE_MIN, _ROA_PLAUSIBLE_MAX) if term == "ROA" else (_ROE_PLAUSIBLE_MIN, _ROE_PLAUSIBLE_MAX)
        if value is not None and not (lo <= value <= hi):
            notes.append(f"implausible value: {value:.2f}% is outside the expected [{lo}%, {hi}%] range for "
                         f"{term} - likely a parsing/scale/sign error, worth a manual check")

        note = "; ".join(notes)
        return {"term": term, "value": value, "matched_label": label, "source_file": source_file,
                "crosscheck_value": crosscheck, "note": note}

    return {
        "roa": build("ROA", "roa_posttax", concall_roa, "roa"),
        "roe": build("ROE", "roe_posttax", concall_roe, "roe"),
    }


def collect_statement_rows(folder, coding, statement, period, verbose=False):
    """Run the account-code extraction for one statement across every .md
    file in `folder`. Returns a list of row dicts (code, name, label_in_doc,
    value, confidence, source_file, page_num). `coding` may be an explicit
    file path, or None to auto-detect the industry category (金控業/金融業/
    保險業) from the filing itself and use its coding workbook."""
    coding = resolve_coding_path(folder=folder, explicit_path=coding)
    code_dict = load_code_dictionary(coding, statement)
    if verbose:
        print(f"Loaded {len(code_dict)} codes for '{statement}' from {coding}")

    rows = []
    for doc_path in sorted(Path(folder).rglob("*.md")):
        results = extract_statement(doc_path, statement, code_dict, period=period, verbose=verbose)
        if results and verbose:
            print(f"[{doc_path.name}] matched {len(results)} code(s)")
        for r in results:
            r["source_file"] = doc_path.name
            rows.append(r)
    return rows


def collect_ratio_rows(folder, bank, coding, verbose=False):
    """Run the ROA/ROE extraction (reported 獲利能力 table, falling back to
    the manual formula). Returns (rows, used_fallback) where rows is a list
    of dicts with period/entity/quarter/roa_posttax/roa_posttax_annualized/
    roe_posttax/roe_posttax_annualized/profit_margin/source_file."""
    entries = find_profitability_entries(folder, verbose=verbose)
    if entries:
        rows = []
        for e in entries:
            rows.append({
                "period": e["period_label"], "entity": e["entity"] or "(this filing)",
                "quarter": e["quarter_num"],
                "roa_posttax": e["roa_posttax"], "roa_posttax_annualized": annualize(e["roa_posttax"], e["quarter_num"]),
                "roe_posttax": e["roe_posttax"], "roe_posttax_annualized": annualize(e["roe_posttax"], e["quarter_num"]),
                "profit_margin": e["profit_margin"], "source_file": e["source_file"],
            })
        return rows, False

    if verbose:
        print("No 獲利能力 disclosure table found; falling back to manual formula.")
    try:
        result = compute_ratios(folder, bank, coding, verbose=verbose)
    except RuntimeError as e:
        print(f"NOTE: manual ROA/ROE fallback also failed - {e}")
        return [], True
    row = {
        "period": None, "entity": "(approximated)", "quarter": result["quarter_num"],
        "roa_posttax": None, "roa_posttax_annualized": result["roa"],
        "roe_posttax": None, "roe_posttax_annualized": result["roe"],
        "profit_margin": None, "source_file": None,
    }
    return [row], True


RATIO_COLUMNS = ["roa_posttax", "roa_posttax_annualized", "roe_posttax", "roe_posttax_annualized", "profit_margin"]
# 'page_num' deliberately excluded: markdown has no page concept, so that
# row field is always None (see build_raw_lines) - the meaningful page
# identifier is the 'page' column derived from source_file via page_num().
STATEMENT_COLUMNS = ["code", "name", "label_in_doc", "value", "confidence"]


def print_statement_rows(statement, rows):
    print(f"\n=== {statement} ===")
    if not rows:
        print("No matching codes found in any file.")
        return
    for r in rows:
        mark = "" if r["confidence"] == "high" else "*"
        print(f"{r['code']}\t{r['name']}\t{format_value(r['value'])}{mark}\t({page_num(r['source_file'])})")


def print_ratio_rows(rows, used_fallback):
    print("\n=== ratios ===")
    if used_fallback and not rows:
        return  # collect_ratio_rows already printed why both the reported table and the fallback failed
    if used_fallback:
        print("NOTE: no reported 獲利能力 table found in this folder - the figures below are "
              "an approximation computed from balance sheet / income statement codes, not a "
              "company-reported number.")
        r = rows[0]
        print(f"ROA(稅後年化, approximated)\t{r['roa_posttax_annualized']:.2%}")
        print(f"ROE(稅後年化, approximated)\t{r['roe_posttax_annualized']:.2%}")
        return
    print("Reported profitability (獲利能力), directly from filing - 稅後 (posttax) only:")
    for r in rows:
        metrics_str = " ".join(f"{c}={format_pct(r[c])}" for c in RATIO_COLUMNS)
        print(f"{r['period']}\t{r['entity']}\t{metrics_str}\t({page_num(r['source_file'])})")


def write_statement_csv(folder, statement, rows):
    out_path = Path(folder) / f"{statement}_export.csv"
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["code", "name", "label_in_doc", "value", "confidence", "page"])
        for r in rows:
            writer.writerow([r["code"], r["name"], r["label_in_doc"], format_value(r["value"]),
                              r["confidence"], page_num(r["source_file"])])
    return out_path


def write_ratio_csv(folder, rows, used_fallback):
    out_path = Path(folder) / "profitability_export.csv"
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if used_fallback and not rows:
            writer.writerow(["error", "no reported 獲利能力 table and manual fallback also failed - see console output"])
        elif used_fallback:
            writer.writerow(["metric", "value_annualized_approximated"])
            r = rows[0]
            writer.writerow(["ROA(稅後年化)", f"{r['roa_posttax_annualized']:.2%}"])
            writer.writerow(["ROE(稅後年化)", f"{r['roe_posttax_annualized']:.2%}"])
        else:
            writer.writerow(["period", "entity", "quarter"] + RATIO_COLUMNS + ["page"])
            for r in rows:
                writer.writerow([r["period"], r["entity"], r["quarter"]] +
                                 [format_pct(r[c]) for c in RATIO_COLUMNS] + [page_num(r["source_file"])])
    return out_path


def write_combined_csv(folder, statement_rows_by_name, ratio_rows, used_fallback):
    """Write one combined CSV covering all three statements plus ratios, with
    a leading 'section' column and blank cells for columns that don't apply
    to a given row's section."""
    out_path = Path(folder) / "combined_export.csv"
    header = ["section"] + STATEMENT_COLUMNS + ["page", "period", "entity", "quarter"] + RATIO_COLUMNS
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for statement, rows in statement_rows_by_name.items():
            for r in rows:
                writer.writerow([statement, r["code"], r["name"], r["label_in_doc"], format_value(r["value"]),
                                  r["confidence"], page_num(r["source_file"]), "", "", ""] + [""] * len(RATIO_COLUMNS))
        for r in ratio_rows:
            blank_statement_cols = [""] * len(STATEMENT_COLUMNS)
            if used_fallback:
                # The fallback row carries BOTH annualized figures - see
                # collect_ratio_rows, and both print_ratio_rows and
                # write_ratio_csv emit both. This branch used to emit one
                # value followed by four blanks, which lost ROE entirely and
                # put ROA in the roa_posttax column instead of the annualized
                # one. The manual formula returns plain ratios, so these two
                # format as :.2% rather than through format_pct.
                def fallback_pct(value):
                    return f"{value:.2%}" if value is not None else "N/A"

                ratio_vals = ["",                                        # roa_posttax
                              fallback_pct(r["roa_posttax_annualized"]),
                              "",                                        # roe_posttax
                              fallback_pct(r["roe_posttax_annualized"]),
                              ""]                                        # profit_margin
            else:
                ratio_vals = [format_pct(r[c]) for c in RATIO_COLUMNS]
            writer.writerow(["ratios"] + blank_statement_cols + [page_num(r["source_file"]), r["period"], r["entity"], r["quarter"]] + ratio_vals)
    return out_path


# ---------------------------------------------------------------------------
# Curated per-bank summary export ("summary" mode): a fixed set of specific
# codes plus two composite/derived terms, with bank-specific variations.
#
# Unlike the whole-statement dumps above, these codes are matched directly
# against a document's rows regardless of whether they're present in a
# coding dictionary - most of them aren't. This is expected - the list spans
# concepts that not every bank reports under the same code, which is also
# why some entries need a bank-specific override (e.g. 國泰 uses 63000
# instead of 64000) or an entirely different composite formula. No
# statement-section restriction is applied, since a given code isn't
# guaranteed to live in the same statement across all 4 banks - the whole
# document is searched.
#
# A code not found in a given filing - or an entire composite term, if any
# one of its component codes is missing - shows as N/A rather than the row
# being omitted, so every expected line always appears.
# ---------------------------------------------------------------------------

# Ordered summary layout: the fixed sequence of lines "summary" mode always
# outputs, each with a CANONICAL display term (not the document's own
# wording, which varies across banks - e.g. 資產合計/資產總計/資產 all become
# 總資產) - a deliberate departure from this mode's earlier "always use the
# document's own label" rule, per explicit instruction to standardize
# output terms for cross-bank comparability. The document's own label is
# still kept (see collect_summary_rows' "matched_label" field) so nothing
# is lost, just no longer what's shown as "term".
# is_cost=True marks a line that should DISPLAY AS POSITIVE for a genuine
# cost (see apply_cost_sign) - a real filing stores an expense as a
# NEGATIVE number in this net-income-walk table style, so a positive
# is_cost value after conversion means "this much cost", and a NEGATIVE
# is_cost value means "this was actually a benefit/reversal, not a cost".
SUMMARY_LAYOUT = [
    {"kind": "code", "code": "10000", "term": "總資產", "is_cost": False},
    {"kind": "code", "code": "4xxxx", "term": "淨收益", "is_cost": False},
    {"kind": "code", "code": "49010", "term": "利息淨收益", "is_cost": False},
    {"kind": "code", "code": "49100", "term": "手續費淨收益", "is_cost": False},
    {"kind": "composite", "name": "評價及已實現", "term": "評價及已實現", "is_cost": False},
    {"kind": "composite", "name": "其他非利息收益", "term": "其他非利息收益", "is_cost": False},
    {"kind": "code", "code": "58400", "term": "營業費用", "is_cost": True},
    {"kind": "code", "code": "58500", "term": "員工福利費用", "is_cost": True},
    {"kind": "code", "code": "59000", "term": "折舊及攤銷費用", "is_cost": True},
    {"kind": "code", "code": "59500", "term": "其他費用", "is_cost": True},
    {"kind": "code", "code": "58200", "term": "呆帳提存(收回)", "is_cost": True},
    {"kind": "code", "code": "61001", "term": "稅前淨利", "is_cost": False},
    {"kind": "code", "code": "61003", "term": "所得稅費用", "is_cost": True},
    {"kind": "code", "code": "64000", "term": "稅後淨利", "is_cost": False},
    # 活存比 has no account code (it's a disclosed ratio, not a coded
    # balance-sheet/income-statement line) - matched purely by label text,
    # like SUMMARY_LABEL_FALLBACKS' subtotal rows (see find_value_by_label),
    # not through build_code_index's code-keyed batch pass. None of the 4
    # banks' filings checked so far disclose it, so this exports N/A for
    # now - kept as a row per explicit instruction, rather than only adding
    # it once a filing with the label actually turns up.
    # "活期性存款比率" is the real wording confirmed in the quarterly
    # summarized fin-report disclosures (see collect_summary_rows_finsum) -
    # kept alongside the earlier guessed wordings in case a full annual/
    # quarterly filing ever uses different phrasing for the same ratio.
    {"kind": "label", "label_aliases": ["活期性存款比率", "活存性存款比率", "活存比"],
     "term": "活存比", "is_percent": True},
]


# SUMMARY_LAYOUT is bank-shaped, and its codes are matched RAW against the
# document - summary mode never loads an industry coding dictionary (see
# collect_summary_rows), so nothing about a code carries its own industry
# with it. The same number means a different account under a different
# scheme: 58200 is 呆帳提存 under 金融業 but an insurance cost line under
# 保險業 (see INDUSTRY_CODING_FILES, which documents that difference).
#
# Applying this layout to a filing it wasn't built for therefore does NOT
# fail - it relabels a real, correctly-parsed number. Confirmed on a 國泰
# 人壽 filing: industry detected as 保險業 correctly, then '保險成本' was
# emitted under the canonical term '呆帳提存(收回)' with its sign flipped by
# apply_cost_sign. matched_label still held '保險成本', but term is the field
# the CSV/Excel exports and _MERGED_TERM_ORDER key on, and standardizing
# term for cross-entity comparison is this layout's whole purpose.
#
# So an industry with no layout of its own is REFUSED, not defaulted.
INDUSTRY_SUMMARY_LAYOUTS = {
    "金融業": SUMMARY_LAYOUT,
    # 金控業 shares the bank layout: the per-bank override tables were built
    # around FHC-consolidated scope in the first place (see
    # SUMMARY_CODE_OVERRIDES_FINSUM's note on 63000/64000), so these codes
    # are already exercised against FHC-scope rows. Not independently
    # re-verified against a 金控's own filing - if one turns up whose rows
    # don't line up, give it its own layout rather than widening this list.
    "金控業": SUMMARY_LAYOUT,
    # 保險業 deliberately absent. A layout for it needs the 保險業 scheme's
    # own codes for 保費收入/保險給付/etc., which nothing in this repo
    # establishes - guessing them would produce exactly the mislabelling
    # above, only harder to spot.
}


def summary_layout_error(industry):
    """Message for a filing whose industry has no summary layout."""
    return (f"summary mode has no layout for industry {industry!r} - it currently covers "
            f"{', '.join(INDUSTRY_SUMMARY_LAYOUTS)}. Its account codes are matched raw "
            f"against the document, so running it on another industry's scheme relabels "
            f"real numbers instead of failing (58200 is 呆帳提存 for a bank but an "
            f"insurance cost line under 保險業). Use a per-statement mode "
            f"(balance_sheet/income_statement/cash_flow), which resolves codes through "
            f"that industry's own coding dictionary.")


def apply_cost_sign(value, matched_label, is_cost):
    """Flip sign so a cost line displays positive (see SUMMARY_LAYOUT's
    is_cost docs). Skipped when the document's OWN label already carries a
    '減' ('less:'/deduct) prefix (confirmed in a real 中信 filing:
    '減：所得稅費用' printed as a POSITIVE number) - that convention already
    stores the value in "amount to subtract" form, i.e. already
    cost-positive, so flipping it again would wrongly turn a normal expense
    into an apparent benefit."""
    if value is None or not is_cost:
        return value
    # A '減：' PREFIX, not a bare '減' anywhere in the label: 減損 (impairment)
    # and 減資 are ordinary line items that merely start with the character,
    # and treating them as already-deducted left a normal expense negative,
    # displaying as if it were a reversal.
    if matched_label and matched_label.strip().startswith(("減：", "減:")):
        return value
    return -value

# Fallback: when a SUMMARY_LAYOUT code isn't found under its own (possibly
# bank-overridden) number, some filings' conversions leave that ROW's own
# leading code cell blank (confirmed in a real 中信 individual filing - the
# subtotal/aggregate lines simply have no code at all), or the true code
# differs from what SUMMARY_CODE_OVERRIDES assumes (confirmed in a real 國泰
# individual filing - "本期淨利" sits at code 61000, not the FHC-consolidated-
# scope 63000/64000 the override was built around). Both cases are really
# the same underlying problem: the CODE is unreliable for these particular
# subtotal/aggregate lines specifically, but the LABEL text is not - these
# lines use one of a small, consistent set of phrasings across all 4 banks'
# filings observed so far. Keyed by the code's SUMMARY_LAYOUT/override slot
# (i.e. the value actually passed to find_code_value), not by whichever
# literal code happens to hold it in a given filing.
SUMMARY_LABEL_FALLBACKS = {
    "10000": ["資產總計", "資產合計"],
    # 淨收益合計 confirmed in a real 第一銀行 114Q4 individual filing - same
    # line, same position in the net-income walk, just worded with the
    # 合計 suffix. Exact whole-cell matching (see find_value_by_label) keeps
    # this from colliding with anything.
    "4xxxx": ["淨收益", "淨收益合計"],
    "49010": ["利息淨收益合計", "利息淨收益"],
    "58400": ["營業費用合計"],
    "61001": ["稅前淨利", "繼續營業單位稅前淨利"],
    "64000": ["本期稅後淨利", "本期淨利", "本年度淨利"],
    "63000": ["本期稅後淨利", "本期淨利", "本年度淨利"],
    "30000": ["權益總計", "權益合計"],
    # 20000 = 負債合計/負債總計 (total liabilities) - not read anywhere
    # currently (was only for the 資產=負債+權益 invariant-check row, since
    # removed from the summary output); kept here, unused, in case that
    # check is wanted again later.
    "20000": ["負債合計", "負債總計"],
}

# Last resort after both the code match and the label fallback fail: rebuild
# the line from its own components. Only for lines a filing can legitimately
# omit entirely - 兆豐 and 第一's real 114Q4 individual filings print 營業費用
# as a section HEADER with no amounts at all, then the three component rows,
# then go straight to 稅前淨利. There is no row to match, by code or by label.
#
# This is arithmetic, not a guess, and it was checked both ways before being
# added: on the four 114Q4 filings that DO print 58400 (台新/新光/永豐/華南)
# the sum of these three reproduces the printed total to the exact dollar,
# and on the two that don't it closes the filing's own
# 淨收益 - 呆帳 - 營業費用 = 稅前淨利 walk exactly. The derived row still says
# so in its note - a figure the filing itself never states must not be
# indistinguishable from one it does.
#
# ponytail: one entry, consulted only on a miss. If a filing ever turns up
# with a FOURTH opex component, this silently understates - which is why the
# note names the components it actually summed.
SUMMARY_CODE_DERIVATIONS = {
    "58400": ["58500", "59000", "59500"],
}

# ---------------------------------------------------------------------------
# Per-entity profiles.
#
# Everything keyed by a reporting entity lives here. It used to live in five
# separate tables (BANKS, BANK_NAME_ALIASES, SUMMARY_CODE_OVERRIDES,
# SUMMARY_CODE_OVERRIDES_FINSUM, COMPOSITE_TERMS) plus
# callfinder.PRIMARY_BANK_ENTITIES - six edits to add one entity, with nothing
# checking they all happened. Every one of those names still exists, derived
# below; only their source moved. _validate_profiles() then makes an
# incomplete entity an import-time error instead of a silent N/A at run time.
#
# FIELDS
#
# industries - which coding schemes this entity's filings are written under.
#   This is what makes the entity axis safe to open up. Aliases have to be
#   short forms (an earnings-call deck's cover says "玉山金控", never a
#   registered name), and short forms collide inside a group: "國泰人壽保險股
#   份有限公司" contains "國泰", so an insurer resolved to its sibling BANK
#   and inherited that bank's overrides and composites. Scoping detection to
#   the filing's own industry removes that by construction instead of hoping
#   the aliases stay distinct. Both bank and FHC schemes are listed for every
#   entity here because both document types are in scope for these groups -
#   the override tables were built around FHC-consolidated scope in the first
#   place (see code_overrides_finsum).
#
# aliases - alternate/full names, used both to normalize a --bank value typed
#   as the full name (e.g. "台北富邦銀行" for 北富銀) and to auto-detect the
#   entity from a filing's own text when --bank isn't given. "臺"/"台" are
#   both included since either can appear in a filing.
#
# primary_entities - the PRIMARY bank subsidiary each earnings-call deck is
#   about. A financial-holding deck also reports OTHER bank subsidiaries
#   (富邦華一銀行 in RMB, 富邦銀行(香港) in HKD, ...) whose tables use
#   identical row labels (存放比, 總放款, 營業費用) - callfinder's
#   _BANK_LABEL_HINT alone can't tell them apart, since every one of them
#   contains "銀行". Matching a figure from the wrong subsidiary is worse than
#   reporting nothing: in a real 富邦 deck it produced 存放比 72.17% and
#   放款餘額 81,769 from the mainland-China subsidiary's RMB-denominated table
#   instead of 台北富邦銀行's own figures. English/romanized forms are
#   included because these decks label appendix tables in English ("E.SUN
#   Bank's Income Statement"); without them such a heading matches
#   _ENTITY_NAME_RE ("Bank") but not the primary alias, and the bank's own
#   appendix would be rejected as if it were another company's.
#
# code_overrides - per-entity overrides for individual SUMMARY_LAYOUT code
#   entries (code -> replacement).
#
# code_overrides_finsum - same idea for the quarterly SUMMARIZED fin-report
#   disclosure (see collect_summary_rows_finsum), a DIFFERENT and shorter
#   document than the full individual filing, confirmed to use a different
#   code scheme for 稅後淨利 in real 114Q4 filings: 中信 prints no code at all
#   for that row (label-only - already covered by
#   SUMMARY_LABEL_FALLBACKS["64000"], no override needed), 北富銀 prints 64000
#   directly (no override needed either), but 國泰 prints 61000 - NOT 63000
#   like the full individual filing needs. Kept as its own field rather than
#   reusing code_overrides since the two document types' schemes genuinely
#   differ for this entity.
#
# composites - this entity's code list for each composite SUMMARY_LAYOUT item.
#   其他非利息收益 was dropped from the default summary output for a while
#   (per an earlier standardized-term-list request) and has been re-added;
#   手續費淨收益 (code 49100, added alongside it) is a plain single-code
#   SUMMARY_LAYOUT item, not a composite, so it has no entry here.
# ---------------------------------------------------------------------------

BANK_PROFILES = {
    "國泰": {
        "industries": ["金融業", "金控業"],
        "aliases": ["國泰"],
        "primary_entities": ["國泰世華", "Cathay United"],
        "code_overrides": {"64000": "63000"},
        "code_overrides_finsum": {"64000": "61000"},
        "composites": {
            "評價及已實現": ["49200", "49310", "49450", "49600"],
            "其他非利息收益": ["49700", "49750", "49800"],
        },
    },
    "中信": {
        "industries": ["金融業", "金控業"],
        "aliases": ["中信", "中國信託"],
        "primary_entities": ["中信銀", "中國信託商業銀行", "中國信託銀行", "CTBC Bank"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            "評價及已實現": ["49200", "49310", "49450", "49600"],
            "其他非利息收益": ["49700", "49750", "49800", "49815", "49899"],
        },
    },
    "北富銀": {
        "industries": ["金融業", "金控業"],
        # "富邦" alone is included so a deck/filing that only ever names the
        # FHC parent ("富邦金控", as the 4Q25 analyst-meeting cover page does)
        # still resolves to this entity; no other profile contains "富邦", so
        # it stays unambiguous.
        "aliases": ["北富銀", "台北富邦銀行", "臺北富邦銀行", "台北富邦", "臺北富邦", "富邦"],
        "primary_entities": ["台北富邦", "臺北富邦", "Taipei Fubon"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            "評價及已實現": ["49200", "49310", "49450", "49600"],
            "其他非利息收益": ["49700", "49750", "49800"],
        },
    },
    "玉山": {
        "industries": ["金融業", "金控業"],
        "aliases": ["玉山"],
        "primary_entities": ["玉山銀", "E.SUN", "ESUN"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            "評價及已實現": ["49200", "49310", "49600"],
            "其他非利息收益": ["49700", "49750", "49899"],
        },
    },
    # The six below were added from real 114Q4 individual filings. Their
    # composites are read off each filing's own 個體綜合損益表 rather than
    # copied from a sibling - the component codes genuinely differ (兆豐 and
    # 第一 print 43100 where everyone else prints 49310; 第一 alone uses 43600
    # and 45000; 華南 uses 47003 for the equity-method line; 新光 has no
    # 除列按攤銷後成本 line at all).
    # primary_entities here are NOT verified against a real earnings-call
    # deck - no deck for these six has been through this tool yet. They only
    # affect con-call extraction; check them against a real deck before
    # trusting con-call output for these entities.
    "兆豐": {
        "industries": ["金融業", "金控業"],
        "aliases": ["兆豐"],
        "primary_entities": ["兆豐銀", "Mega International", "Mega Bank"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            "評價及已實現": ["49200", "43100", "49450", "49600"],
            "其他非利息收益": ["49700", "49750", "49800"],
        },
    },
    "台新": {
        "industries": ["金融業", "金控業"],
        "aliases": ["台新", "臺新"],
        "primary_entities": ["台新銀", "臺新銀", "Taishin"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            "評價及已實現": ["49200", "49310", "49450", "49600"],
            "其他非利息收益": ["49700", "49750", "49800"],
        },
    },
    "新光": {
        "industries": ["金融業", "金控業"],
        "aliases": ["新光"],
        "primary_entities": ["新光銀", "Shin Kong Bank"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            # No 除列按攤銷後成本衡量之金融資產損益 line in this filing at all.
            "評價及已實現": ["49200", "49310", "49600"],
            "其他非利息收益": ["49700", "49815", "49899"],
        },
    },
    "永豐": {
        "industries": ["金融業", "金控業"],
        "aliases": ["永豐"],
        "primary_entities": ["永豐銀", "Bank SinoPac"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            "評價及已實現": ["49200", "49310", "49450", "49600"],
            "其他非利息收益": ["49700", "49750", "49800"],
        },
    },
    # "第一" alone would be a substring of ordinary text (第一階段, 第一季) in
    # every other bank's filing, making this entity a candidate everywhere and
    # turning detect_bank ambiguous for all of them - hence only the longer
    # forms.
    "第一": {
        "industries": ["金融業", "金控業"],
        "aliases": ["第一商業銀行", "第一銀行", "第一金"],
        "primary_entities": ["第一銀", "First Commercial Bank"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            "評價及已實現": ["49200", "43100", "43600", "49600"],
            "其他非利息收益": ["45000", "49750", "49800"],
        },
    },
    "華南": {
        "industries": ["金融業", "金控業"],
        "aliases": ["華南"],
        "primary_entities": ["華南銀", "Hua Nan"],
        "code_overrides": {},
        "code_overrides_finsum": {},
        "composites": {
            "評價及已實現": ["49200", "49310", "49450", "49600"],
            "其他非利息收益": ["49700", "47003", "49899"],
        },
    },
}

_PROFILE_FIELDS = {"industries", "aliases", "primary_entities",
                   "code_overrides", "code_overrides_finsum", "composites"}


def _validate_profiles(profiles):
    """Reject an incomplete or inconsistent entity profile at import time.

    The whole point of consolidating six tables into one is that adding an
    entity can't half-happen any more. A missing composites entry used to
    surface as a single N/A row in one bank's summary - indistinguishable
    from a genuinely undisclosed line, and only in that bank's output."""
    for name, profile in profiles.items():
        missing = _PROFILE_FIELDS - set(profile)
        extra = set(profile) - _PROFILE_FIELDS
        if missing or extra:
            raise ValueError(f"entity profile {name!r} is malformed"
                             + (f"; missing {sorted(missing)}" if missing else "")
                             + (f"; unexpected {sorted(extra)}" if extra else ""))
        if not profile["aliases"]:
            raise ValueError(f"entity profile {name!r} has no aliases, so it can never be detected")
        for industry in profile["industries"]:
            if industry not in INDUSTRY_CODING_FILES:
                raise ValueError(f"entity profile {name!r} names industry {industry!r}, which is not "
                                 f"one of {list(INDUSTRY_CODING_FILES)}")
        # Every composite item in a layout this entity's filings are read
        # under needs a code list here - that is the check the six separate
        # tables could not perform.
        for industry in profile["industries"]:
            for item in INDUSTRY_SUMMARY_LAYOUTS.get(industry) or []:
                if item["kind"] == "composite" and item["name"] not in profile["composites"]:
                    raise ValueError(f"entity profile {name!r} is read under {industry!r}, whose summary "
                                     f"layout has composite item {item['name']!r}, but defines no codes "
                                     f"for it")


_validate_profiles(BANK_PROFILES)

# Derived views. These are the names the rest of the codebase (and the tests)
# already use; BANK_PROFILES is simply where they now come from.
BANKS = list(BANK_PROFILES)
BANK_NAME_ALIASES = {name: p["aliases"] for name, p in BANK_PROFILES.items()}
SUMMARY_CODE_OVERRIDES = {name: p["code_overrides"] for name, p in BANK_PROFILES.items()}
SUMMARY_CODE_OVERRIDES_FINSUM = {name: p["code_overrides_finsum"] for name, p in BANK_PROFILES.items()}


def _invert_composites(profiles):
    """term -> {entity: codes}, the shape collect_summary_rows reads."""
    inverted = {}
    for name, profile in profiles.items():
        for term, codes in profile["composites"].items():
            inverted.setdefault(term, {})[name] = codes
    return inverted


COMPOSITE_TERMS = _invert_composites(BANK_PROFILES)


def resolve_bank_name(name):
    """Normalize a --bank value to its canonical BANKS entry, accepting
    either the short form or any alias in BANK_NAME_ALIASES. Raises
    ValueError with the accepted names if nothing matches."""
    for canonical, aliases in BANK_NAME_ALIASES.items():
        if name == canonical or name in aliases:
            return canonical
    accepted = ", ".join(f"{b} ({'/'.join(BANK_NAME_ALIASES[b])})" for b in BANKS)
    raise ValueError(f"Unrecognized bank '{name}'. Accepted: {accepted}")


def bank_candidates(folder, industry=None):
    """Every entity in BANK_PROFILES whose alias appears in the first few .md
    files of `folder` (sorted by name - typically the cover/first pages), in
    BANKS order. Empty list if no file exists or nothing matches.

    Restricted to entities whose `industries` include the filing's own, so a
    group's insurer can't be resolved as its sibling bank purely because the
    short alias is a substring of both. `industry` may be passed by a caller
    that already resolved it; None means detect it. When the industry can't
    be established at all, every entity stays a candidate - an earnings-call
    deck carries no registered name and so never resolves an industry, and
    narrowing there would break con-call detection outright.

    Reads the same first-5 window detect_industry_category does. Reading only
    paths[0] meant a filing whose cover page didn't carry the bank's name -
    the exact case detect_industry_category was widened to 5 files for -
    detected its industry fine but failed on the bank, and runfinder then
    skipped the whole folder."""
    paths = sorted(Path(folder).rglob("*.md"))[:5]
    if not paths:
        return []
    text = "".join(p.read_text(encoding="utf-8", errors="ignore") for p in paths)
    if industry is None:
        industry = detect_industry_category(folder)
    return [name for name, profile in BANK_PROFILES.items()
            if (industry is None or industry in profile["industries"])
            and any(alias in text for alias in profile["aliases"])]


def detect_bank(folder):
    """The single bank `folder` belongs to, or None when that can't be
    established - either because nothing matched, or because MORE THAN ONE
    bank did.

    This used to return the first match in BANK_NAME_ALIASES order, which is
    only safe while no filing can name two of them. That assumption does not
    survive the full set of Taiwanese banks: the aliases are short forms
    (they have to be - an earnings-call deck's cover says '玉山金控', never
    the legal name, so detect_industry_category's full-legal-name approach
    can't be borrowed here), many are substrings of each other, and a filing
    naming a peer in a related-party or interbank note is ordinary rather
    than exceptional. Under those conditions first-match-wins doesn't
    degrade to N/A - it silently picks the wrong bank, applies that bank's
    COMPOSITE_TERMS and SUMMARY_CODE_OVERRIDES, and produces a full set of
    plausible-looking wrong numbers.

    Ambiguity is therefore refused rather than guessed. Callers already have
    a path for "couldn't detect" (acctfinder asks for --bank, runfinder skips
    the folder and says so), so this needs no new control flow - only a
    message that distinguishes the two cases, which is what bank_candidates
    is exposed for."""
    matched = bank_candidates(folder)
    return matched[0] if len(matched) == 1 else None


def bank_detection_message(folder):
    """The --bank guidance to show when detect_bank returned None. The two
    reasons need different actions from whoever reads it: nothing matched
    means this entity may not be supported at all, while several matched
    means it IS supported and only needs disambiguating - so they must not
    collapse into one message."""
    found = bank_candidates(folder)
    if found:
        return ("Several banks are named in this filing's first pages (" + ", ".join(found)
                + ") - pass --bank explicitly to say which one is the reporting entity")
    # A filing whose industry IS resolvable but which no profile covers is a
    # third case: --bank can't help, because every accepted value would then
    # be refused again by the layout guard. Saying "couldn't auto-detect"
    # here sends the reader down a dead end.
    industry = detect_industry_category(folder)
    supported = {i for p in BANK_PROFILES.values() for i in p["industries"]}
    if industry and industry not in supported:
        return (f"This filing reads as {industry}, and no supported entity files under that scheme "
                f"yet - the entities in BANK_PROFILES cover {', '.join(sorted(supported))}")
    return ("Couldn't auto-detect the bank from the filing's text - pass --bank "
            "explicitly (choices: " + ", ".join(BANKS) + ")")


# The three folder-wide lookups these extractors sit on moved to
# core/lookup.py - both compute_ratios and collect_summary_rows need
# them and neither owns them. Re-exported above, so `acctfinder.X`
# still resolves for callers and for tests that monkeypatch it.


def collect_summary_rows(folder, bank, period=1, verbose=False, coding=None,
                          concall_roa=None, concall_roe=None, overrides_table=None,
                          industry=None):
    """Build the curated summary, in SUMMARY_LAYOUT's fixed order, plus a
    trailing ROA and ROE row (see collect_roa_roe). Returns a list of
    {term, value, matched_label, source_file, is_percent, crosscheck_value,
    note}:
      - term: the CANONICAL display name from SUMMARY_LAYOUT (not the
        document's own wording - see SUMMARY_LAYOUT's docstring).
      - matched_label: the document's own label for that row (or the term
        itself for a composite, which has no single source row) - kept so
        the original wording isn't lost even though it's no longer "term".
      - value: sign-adjusted per is_cost (see apply_cost_sign); None
        (displayed as N/A) rather than the row being omitted when a code
        genuinely can't be found, so every expected line always appears.
      - is_percent: True only for ROA/ROE (already a % rate, not a NT$
        amount) - display code uses this to choose format_pct vs
        format_value rather than assuming based on term name.
      - crosscheck_value: ROA/ROE only, None for every other row (see
        collect_roa_roe - the manual-formula cross-check).
      - note: "" for a row that resolved cleanly. Set on ROA/ROE when the
        cross-check diverges or the value is implausible (collect_roa_roe),
        AND on every N/A row, saying WHY it's N/A - missing code, no
        matching label, no composite formula for this bank. That reason is
        also printed under -v, but the note is the only copy that reaches
        the csv/excel exports, which is where the distinction between "not
        disclosed" and "we failed to read this filing" actually matters.
    concall_roa/concall_roe: an earnings-call deck's own reported ROA/ROE
    (looked up by the caller via callfinder.py, since this module can't
    import it - see collect_roa_roe), used as a fallback when this fin
    folder has no reported 獲利能力 disclosure table of its own.
    overrides_table: which per-bank code-override table to apply (defaults
    to SUMMARY_CODE_OVERRIDES, for full individual filings) - passed as
    SUMMARY_CODE_OVERRIDES_FINSUM by collect_summary_rows_finsum(), since
    the quarterly summarized disclosure uses a different code scheme for
    at least one bank (see that table's docs). Every other piece of logic
    here (SUMMARY_LAYOUT, COMPOSITE_TERMS, label fallbacks, ROA/ROE, CIR)
    is identical between the two document types - only the code-resolution
    step differs, which is why this is a shared function with a swappable
    table rather than a separate parallel implementation.
    """
    # Resolved here rather than at the call sites: this is the only place
    # that knows a layout is about to be applied, so a future caller can't
    # forget the check. industry may be passed in when the caller already
    # resolved it; None means detect it from the filing.
    if industry is None:
        industry = detect_industry_category(folder)
    layout = INDUSTRY_SUMMARY_LAYOUTS.get(industry)
    if layout is None:
        raise ValueError(summary_layout_error(industry))

    overrides = (overrides_table if overrides_table is not None else SUMMARY_CODE_OVERRIDES).get(bank, {})

    # Gather every code this run will need up front, so build_code_index can
    # resolve all of them in one shared pass over the folder's files instead
    # of one pass per code (see build_code_index) - both the "code"-kind
    # items (after per-bank override resolution) and every composite item's
    # component codes.
    needed_codes = set()
    label_fallbacks = {}
    for item in layout:
        if item["kind"] == "code":
            code = overrides.get(item["code"], item["code"])
            needed_codes.add(code)
            # Fetched whether or not they're needed - they're layout items in
            # their own right today, but this must not depend on that.
            needed_codes.update(SUMMARY_CODE_DERIVATIONS.get(code, []))
            fallback = SUMMARY_LABEL_FALLBACKS.get(item["code"])
            if fallback:
                label_fallbacks[code] = fallback
        elif item["kind"] == "composite":
            needed_codes.update(COMPOSITE_TERMS[item["name"]].get(bank) or [])
        # "label" kind (e.g. 活存比) has no account code at all - resolved
        # separately via find_value_by_label in the row-building loop below,
        # not part of this code-keyed batch pass.

    index = build_code_index(folder, needed_codes, label_fallbacks=label_fallbacks,
                              period=period, verbose=verbose)

    rows = []
    for item in layout:
        if item["kind"] == "code":
            code = overrides.get(item["code"], item["code"])
            found = index.get(code)
            note = ""
            if found is None and code in SUMMARY_CODE_DERIVATIONS:
                parts = [index.get(c) for c in SUMMARY_CODE_DERIVATIONS[code]]
                if all(p is not None for p in parts):
                    labels = [p[0] for p in parts]
                    found = ("+".join(labels), sum(p[1] for p in parts), parts[0][2])
                    note = (f"derived: this filing states no {item['term']} total, "
                            f"so this is {' + '.join(labels)}")
                    if verbose:
                        print(f"[derived] code {code} ({item['term']}) = {' + '.join(labels)}")
            if found is None:
                # Same string to the console and to the row's note: the note is
                # the only place this reason survives into the exported csv/
                # excel, where an unreadable filing and an undisclosed line
                # otherwise look identical (see summary_coverage_warning).
                reason = f"code {code} ({item['term']}) not found in any file"
                if verbose:
                    print(f"[N/A] {reason}")
                rows.append({"term": item["term"], "value": None, "matched_label": None, "source_file": None,
                             "is_percent": False, "crosscheck_value": None, "note": reason})
                continue
            matched_label, value, source_file = found
            value = apply_cost_sign(value, matched_label, item["is_cost"])
            rows.append({"term": item["term"], "value": value,
                         "matched_label": matched_label, "source_file": source_file,
                         "is_percent": False, "crosscheck_value": None, "note": note})
            continue

        if item["kind"] == "label":
            term_name = item["term"]
            found = find_value_by_label(folder, item["label_aliases"], period=period, verbose=verbose)
            if found is None:
                reason = f"'{term_name}' - none of {item['label_aliases']} found in any file"
                if verbose:
                    print(f"[N/A] {reason}")
                rows.append({"term": term_name, "value": None, "matched_label": None, "source_file": None,
                             "is_percent": item.get("is_percent", False), "crosscheck_value": None,
                             "note": reason})
                continue
            matched_label, value, source_file = found
            rows.append({"term": term_name, "value": value,
                         "matched_label": matched_label, "source_file": source_file,
                         "is_percent": item.get("is_percent", False), "crosscheck_value": None, "note": ""})
            continue

        # composite
        term_name = item["term"]
        codes = COMPOSITE_TERMS[item["name"]].get(bank)
        if codes is None:
            reason = f"'{term_name}' has no formula defined for bank '{bank}'"
            if verbose:
                print(f"[N/A] {reason}")
            rows.append({"term": term_name, "value": None, "matched_label": None, "source_file": None,
                         "is_percent": False, "crosscheck_value": None, "note": reason})
            continue
        component_values, component_files, missing = [], [], None
        for code in codes:
            found = index.get(code)
            if found is None:
                missing = f"'{term_name}': component code {code} not found"
                if verbose:
                    print(f"[N/A] {missing}")
                break
            _label, value, source_file = found
            component_values.append(value)
            component_files.append(source_file)
        if missing:
            rows.append({"term": term_name, "value": None, "matched_label": None, "source_file": None,
                         "is_percent": False, "crosscheck_value": None, "note": missing})
            continue
        total = sum(component_values)
        total = apply_cost_sign(total, None, item["is_cost"])
        rows.append({"term": term_name, "value": total,
                     "matched_label": term_name, "source_file": component_files[0],
                     "is_percent": False, "crosscheck_value": None, "note": ""})

    # --- CIR: abs(營業費用) / 淨收益, direct from THIS filing, no crosscheck.
    # Previously computed on the con-call side from that deck's own
    # 營業費用/營業收入 table, but that figure turned out to be a different
    # scope from this fin_report's individual-entity 58400/4xxxx codes (e.g.
    # a real 中信 4Q25 con-call page's revenue/opex were roughly an order of
    # magnitude off fin_report's, not a rounding-level gap) - moved here per
    # explicit instruction, no extra folder scan needed.
    #
    # Read off the ROWS just built, not the raw code index: those two rows
    # have already been through the label fallback and SUMMARY_CODE_DERIVATIONS,
    # so a filing that states no 營業費用 total still gets a CIR instead of
    # this silently disagreeing with the 營業費用 line printed right above it.
    # abs() makes the is_cost sign flip on the 營業費用 row irrelevant here.
    by_term = {r["term"]: r["value"] for r in rows}
    netrev_value, opex_value = by_term.get("淨收益"), by_term.get("營業費用")
    cir_value, cir_note = None, ""
    if netrev_value is not None and opex_value is not None:
        if netrev_value:
            cir_value = abs(opex_value) / netrev_value * 100
        else:
            cir_note = "CIR undefined: 淨收益 is zero"
    else:
        absent = [t for t, v in (("淨收益", netrev_value), ("營業費用", opex_value)) if v is None]
        cir_note = f"CIR needs {' and '.join(absent)}, which came back N/A above"
    rows.append({"term": "CIR", "value": cir_value, "matched_label": "abs(營業費用) / 淨收益",
                 "source_file": None, "is_percent": True, "crosscheck_value": None, "note": cir_note})

    roa_roe = collect_roa_roe(folder, bank, coding=coding, concall_roa=concall_roa,
                               concall_roe=concall_roe, verbose=verbose)
    for key, term in (("roa", "ROA"), ("roe", "ROE")):
        r = roa_roe[key]
        if r is None:
            # None means all three sources in collect_roa_roe's priority order
            # came up empty - naming them is what tells a reader whether to go
            # looking for a 獲利能力 table or for the missing balance-sheet
            # codes the manual formula needs.
            rows.append({"term": term, "value": None, "matched_label": None, "source_file": None,
                         "is_percent": True, "crosscheck_value": None,
                         "note": f"no {term}: filing discloses no 獲利能力 table, no con-call figure "
                                 f"was supplied, and the manual formula wasn't derivable"})
        else:
            rows.append({"term": r["term"], "value": r["value"], "matched_label": r["matched_label"],
                         "source_file": r["source_file"], "is_percent": True,
                         "crosscheck_value": r["crosscheck_value"], "note": r["note"]})

    return rows


def collect_summary_rows_finsum(folder, bank, period=1, verbose=False, coding=None,
                                 concall_roa=None, concall_roe=None, industry=None):
    """Curated summary for a bank's quarterly SUMMARIZED fin-report
    disclosure (依「公開發行銀行財務報告編製準則」第三十二條規定網站揭露財務
    業務資訊) - a much shorter document (~10-15 pages) than the full
    individual filing (~150-280), but confirmed (114Q4 CTBC/北富銀/國泰) to
    use the exact same table shapes, code conventions, and section content
    for every SUMMARY_LAYOUT item - including CTBC's dual-column balance
    sheet (already handled by build_raw_lines' _split_dual_column_tables)
    and a real 活期性存款比率 disclosure (see SUMMARY_LAYOUT's 活存比 entry).
    The only genuine difference found is 稅後淨利's code varying by bank in
    THIS document type specifically (see SUMMARY_CODE_OVERRIDES_FINSUM) -
    everything else is identical, so this is a thin wrapper around
    collect_summary_rows rather than a separate parsing implementation, per
    explicit instruction to reuse the same logic as the full filing.
    Deliberately NOT wired into any automatic/default code path - only
    called once a caller (runfinder.py's folder classifier) has positively
    identified a folder as this document type, never run on a folder just
    because it happens to be a fin_report."""
    return collect_summary_rows(folder, bank, period=period, verbose=verbose, coding=coding,
                                 concall_roa=concall_roa, concall_roe=concall_roe,
                                 overrides_table=SUMMARY_CODE_OVERRIDES_FINSUM,
                                 industry=industry)


# Fraction of N/A rows above which a summary is more likely a failed read
# than a genuinely absent disclosure. Deliberately generous: a handful of
# N/A rows is ordinary (活存比 is N/A for every bank checked so far, and a
# filing that truly doesn't disclose a line is a real result), so this has
# to clear normal variation and only fire on wholesale failure.
_SUMMARY_NA_WARN_RATIO = 0.5


def summary_coverage_warning(rows, folder=None):
    """A one-line warning when most of a summary came back N/A, else None.

    An N/A row still occupies a line like any other (every SUMMARY_LAYOUT
    line always appears, see collect_summary_rows), so a filing whose layout
    this extractor simply failed to read produces output shaped exactly like
    a successful run. Each N/A row now carries its own reason in `note`, but
    that is per-row: it says why THIS line is missing, not that the run as a
    whole went wrong. At four banks the wholesale case was catchable by eye;
    across the whole sector it is not, and the csv/excel export paths don't
    even print the rows for anyone to look at.

    ponytail: one flat ratio over all rows. If it turns out to be noisy, the
    fix is a per-row 'expected N/A' flag in SUMMARY_LAYOUT (活存比 is the
    known case), not a tuned constant - the constant can't tell the two
    kinds of N/A apart no matter what it's set to."""
    if not rows:
        return None
    missing = [r["term"] for r in rows if r["value"] is None]
    if len(missing) <= _SUMMARY_NA_WARN_RATIO * len(rows):
        return None
    where = f" for {folder}" if folder else ""
    return (f"WARNING: {len(missing)} of {len(rows)} summary rows are N/A{where} - "
            f"check this filing's layout is being read correctly ({', '.join(missing)})")


def print_summary_rows(rows):
    print("\n=== summary ===")
    if not rows:
        print("No matching codes found in any file.")
        return
    for r in rows:
        found = r.get("matched_label") or "-"
        value_str = format_maybe_pct(r["value"], r.get("is_percent", False))
        line = f"{r['term']}\t{value_str}\t{found}\t({page_num(r['source_file'])})"
        # Only surface the cross-check value when it's BOTH available AND
        # something was actually flagged about this row (crosscheck_value is
        # populated whenever the manual formula is derivable, whether or not
        # it agrees - showing it unconditionally would defeat the earlier
        # "don't show a cross-check that just confirms the primary value"
        # design; requiring a note too keeps it hidden when everything
        # agrees, and correctly hides it for the invariant-check row below,
        # which always has a note but never a crosscheck_value at all).
        if r.get("crosscheck_value") is not None and r.get("note"):
            line += f"\tcross-check: {format_pct(r['crosscheck_value'])}"
        print(line)
        if r.get("note"):
            print(f"  NOTE: {r['note']}")


def write_summary_csv(folder, rows):
    out_path = Path(folder) / "summary_export.csv"
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["term", "value", "term_found", "page", "crosscheck_value", "note"])
        for r in rows:
            value_str = format_maybe_pct(r["value"], r.get("is_percent", False))
            # Only surface the cross-check value when it's both available and
            # something was flagged - see the matching comment in print_summary_rows.
            crosscheck = (format_pct(r["crosscheck_value"])
                          if r.get("crosscheck_value") is not None and r.get("note") else "")
            writer.writerow([r["term"], value_str, r.get("matched_label") or "",
                              page_num(r["source_file"]), crosscheck, r.get("note") or ""])
    return out_path


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("folder", nargs="?", default=None,
                     help="Folder containing .md files. If omitted, a folder-picker dialog opens.")
    ap.add_argument("statement", choices=list(STATEMENTS) + ["equity_statement", "ratios", "all", "summary"],
                     help="Which statement to extract, 'ratios' for calculated ROA/ROE, "
                          "or 'all' for balance_sheet + income_statement + cash_flow + ratios together")
    ap.add_argument("--coding", default=None,
                     help="Path to an account-coding dictionary .xlsx, overriding auto-detection. "
                          "By default the industry category (金控業/金融業/保險業) is auto-detected "
                          "from the filing's own entity name and its matching bundled coding file is used.")
    ap.add_argument("--industry", choices=list(INDUSTRY_CODING_FILES),
                     help="Force the industry category instead of auto-detecting it from the filing.")
    ap.add_argument("--period", type=int, default=1,
                     help="Which period to extract, counting left-to-right as listed in the "
                          "document (these filings always list the most recent period first). "
                          "1 = most recent (default), 2 = next most recent, etc.")
    ap.add_argument("--export", choices=["csv"], help="Write results to a CSV file instead of stdout")
    ap.add_argument("--bank",
                     help="Bank name for 'summary' (resolves bank-specific code overrides and "
                          "composite-term formulas). Accepts the short form or full name, e.g. "
                          "either '北富銀' or '台北富邦銀行'. If omitted, auto-detected from the "
                          "filing's own text. Choices: " + ", ".join(BANKS))
    ap.add_argument("-v", "--verbose", action="store_true", help="Print per-file/per-table detail")
    args = ap.parse_args()

    if args.period < 1:
        ap.error("--period must be 1 or greater (1 = the most recent period).")

    if args.statement == "equity_statement":
        # Kept in `choices` so --help still lists it and the user gets THIS
        # message rather than "invalid choice". It used to reach
        # load_code_dictionary and surface as a bare ValueError traceback,
        # despite the docstring promising a clear error.
        ap.error(UNSUPPORTED_STATEMENT_MSG)

    if args.folder is None:
        args.folder = pick_folder()
        if args.folder is None:
            ap.error("No folder selected.")

    if args.statement != "summary":
        # 'summary' never touches a coding dictionary (see collect_summary_rows -
        # it matches raw document codes directly, never the dictionary's name),
        # so industry resolution only matters for the other statement modes.
        try:
            args.coding = resolve_coding_path(industry=args.industry, folder=args.folder, explicit_path=args.coding)
        except ValueError as e:
            ap.error(str(e))
        if args.verbose:
            print(f"Using coding file: {args.coding}")

    if args.statement == "summary":
        if args.bank:
            try:
                bank = resolve_bank_name(args.bank)
            except ValueError as e:
                ap.error(str(e))
        else:
            bank = detect_bank(args.folder)
            if bank is None:
                ap.error(bank_detection_message(args.folder))
            if args.verbose:
                print(f"Auto-detected bank: {bank}")
        try:
            rows = collect_summary_rows(args.folder, bank, period=args.period,
                                         verbose=args.verbose, industry=args.industry)
        except ValueError as e:
            ap.error(str(e))
        # Warned here rather than in print_summary_rows: --export csv returns
        # without ever printing the rows, and that is the path where a failed
        # read is least likely to be noticed.
        warning = summary_coverage_warning(rows)
        if warning:
            print(warning)
        if args.export == "csv":
            out_path = write_summary_csv(args.folder, rows)
            print(f"Wrote {len(rows)} row(s) to {out_path}")
            return
        print_summary_rows(rows)
        return

    if args.statement in ("all", "ratios"):
        if args.bank:
            try:
                bank = resolve_bank_name(args.bank)
            except ValueError as e:
                ap.error(str(e))
        else:
            bank = detect_bank(args.folder)
            if bank is None:
                ap.error(bank_detection_message(args.folder))

    if args.statement == "all":
        statement_rows_by_name = {
            s: collect_statement_rows(args.folder, args.coding, s, args.period, args.verbose)
            for s in STATEMENTS
        }
        ratio_rows, used_fallback = collect_ratio_rows(args.folder, bank, args.coding, args.verbose)

        if args.export == "csv":
            out_path = write_combined_csv(args.folder, statement_rows_by_name, ratio_rows, used_fallback)
            total = sum(len(r) for r in statement_rows_by_name.values()) + len(ratio_rows)
            print(f"Wrote {total} row(s) to {out_path}")
            return

        for statement, rows in statement_rows_by_name.items():
            print_statement_rows(statement, rows)
        print_ratio_rows(ratio_rows, used_fallback)
        return

    if args.statement == "ratios":
        rows, used_fallback = collect_ratio_rows(args.folder, bank, args.coding, args.verbose)
        if args.export == "csv":
            out_path = write_ratio_csv(args.folder, rows, used_fallback)
            print(f"Wrote {len(rows)} row(s) to {out_path}")
            return
        print_ratio_rows(rows, used_fallback)
        return

    rows = collect_statement_rows(args.folder, args.coding, args.statement, args.period, args.verbose)
    if args.export == "csv":
        out_path = write_statement_csv(args.folder, args.statement, rows)
        print(f"Wrote {len(rows)} row(s) to {out_path}")
        return
    print_statement_rows(args.statement, rows)


if __name__ == "__main__":
    main()
