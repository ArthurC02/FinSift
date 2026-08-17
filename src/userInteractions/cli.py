"""
Pick up to 2 folders of converted markdown (.md) files, auto-detect which is a
financial report and which is an earnings-call deck, and run the right
extractor's curated summary on each.

Detection uses two signals - exact account codes in a table row's first cell
(structural) and con-call cover markers (textual). CODED ROWS WIN when both
are present: a deck can use a statement's section NAME as an ordinary row
label, which false-positives on a text-only search.
  → docs/knowledge/cli-and-export.md#資料夾怎麼分類

Usage:
    python cli.py [--config con_call_terms.json] [--export csv] [-v]

With no arguments, opens the folder-picker up to twice (cancel the second
dialog if you only have one folder to classify).
"""

import argparse
import importlib
import re
import sys
from pathlib import Path

# `python src/userInteractions/cli.py` puts THIS directory on sys.path, not
# src/, so sibling packages would not resolve. This is the one file users run
# by path, so it bootstraps src/ rather than requiring PYTHONPATH.
_SRC = Path(__file__).resolve().parent.parent
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

import financialReports as fin
import earningsCalls as ec

# Windows consoles often default to cp1252, which can't print CJK output.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

# Repo root is THREE levels up from src/<package>/, NOT two - two levels
# silently points at src/. This has bitten four modules in this repo.
#   → docs/knowledge/cli-and-export.md#子命令為什麼用-peek-argv-而不是-subparsers
_DEFAULT_CONFIG = str(Path(__file__).resolve().parent.parent.parent / "data" / "con_call_terms.json")
_EXCEL_EXPORT_DIR = Path.home() / "Downloads"

_CON_CALL_MARKERS = ["法說會", "法人說明會", "說明會", "說明簡報", "投資人簡報",
                     "投資人關係", "法人電話會議"]


def load_all_codes():
    """Union of every code across all 3 industry coding files and all 3
    statements - only for recognising "this row's leading cell is SOME real
    account code" in classify_folder, never for resolving a specific
    industry's dictionary (that happens later, per folder)."""
    codes = set()
    for path in fin.INDUSTRY_CODING_FILES.values():
        for stmt in fin.STATEMENTS:
            codes.update(fin.load_code_dictionary(path, stmt).keys())
    return codes


def pick_folders(max_count=2):
    """Open the folder-picker up to max_count times, stopping early if the
    user cancels (picking just 1 folder is normal, not an error). Opens in
    Downloads - Windows' native dialog otherwise remembers whatever directory
    some OTHER, unrelated program last used it in."""
    import tkinter as tk
    from tkinter import filedialog
    folders = []
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    downloads = Path.home() / "Downloads"
    initialdir = str(downloads) if downloads.is_dir() else None
    for i in range(max_count):
        title = f"Select folder {i + 1} of up to {max_count} (Cancel if done)"
        folder = filedialog.askdirectory(title=title, initialdir=initialdir)
        if not folder:
            break
        folders.append(folder)
    root.destroy()
    return folders


# Marker + file-count ceiling TOGETHER. The marker alone would also fire on a
# full filing that happens to disclose this ratio somewhere among its 150-280
# files; the summarized disclosure is ~10-15.
#   → docs/knowledge/cli-and-export.md#簡式揭露怎麼認
_FINSUM_MARKER = "活期性存款比率"
_FINSUM_MAX_FILES = 30

# Row order for a merged fin+con sheet. Deliberately spelled out in FULL (not
# "fin terms then leftover con terms") so this list stays the single source of
# truth; merge_fin_and_con_rows falls back to source order only for a term not
# listed here at all.
#   → docs/knowledge/cli-and-export.md#合併輸出的列順序
_MERGED_TERM_ORDER = [
    "總資產", "淨收益", "利息淨收益", "手續費淨收益", "評價及已實現",
    "其他非利息收益", "營業費用", "員工福利費用", "折舊及攤銷費用",
    "其他費用", "呆帳提存(收回)", "稅前淨利", "所得稅費用", "稅後淨利",
    "ROA", "ROE", "活存比", "CIR",
    "NIM", "存放利差", "放款均率", "存款均率",
    "逾放比率", "備抵呆帳/逾期放款",
    "企業放款", "房貸", "個人放款", "信用卡循環",
    "法說會放款餘額合計", "法說會外幣放款",
]


def merge_fin_and_con_rows(fin_rows, con_rows):
    """fin_rows/con_rows: excel_rows lists (see fin_report_rows/con_call_rows)
    from one paired fin_report(+summary) folder and one con_call folder.
    Returns a single reordered list per _MERGED_TERM_ORDER, each term
    appearing exactly once even though it could in principle exist on both
    sides (it never does in practice, but fin_rows is checked first since
    活存比/CIR only live there)."""
    fin_by_term = {r[0]: r for r in fin_rows}
    con_by_term = {r[0]: r for r in con_rows}
    used_fin, used_con = set(), set()
    merged = []
    for term in _MERGED_TERM_ORDER:
        if term in fin_by_term:
            merged.append(fin_by_term[term])
            used_fin.add(term)
            # A same-named con-call row is SUPERSEDED, not still pending -
            # without this it falls through to the trailing extend and the
            # term is emitted twice.
            used_con.add(term)
        elif term in con_by_term:
            merged.append(con_by_term[term])
            used_con.add(term)
    merged.extend(r for r in fin_rows if r[0] not in used_fin)
    merged.extend(r for r in con_rows if r[0] not in used_con)
    return merged


def classify_folder(folder, codes, code_row_threshold=5):
    """Return ('fin_report' | 'fin_report_summary' | 'con_call' | None,
    code_hits, con_hits). code_hits counts table rows whose first cell is
    an exact account code; con_hits counts files containing a con-call
    cover-page marker. 'fin_report_summary' (see _FINSUM_MARKER) routes to
    financialReports.collect_summary_rows_finsum() instead of the full-filing
    collect_summary_rows() - same SUMMARY_LAYOUT items, different per-bank
    code scheme for 稅後淨利 (see SUMMARY_CODE_OVERRIDES_FINSUM)."""
    code_hits = 0
    con_hits = 0
    has_finsum_marker = False
    # rglob, NOT glob - matching every other folder scan in the project. With
    # glob, a folder whose .md files sit in a subdirectory classifies as None
    # and is skipped, while the statement modes read it fine.
    paths = sorted(Path(folder).rglob("*.md"))
    for path in paths:
        text = path.read_text(encoding="utf-8", errors="ignore")
        if any(m in text for m in _CON_CALL_MARKERS):
            con_hits += 1
        if _FINSUM_MARKER in text:
            has_finsum_marker = True
        for line in text.splitlines():
            if not line.startswith("|"):
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if cells and cells[0] in codes:
                code_hits += 1
    if code_hits >= code_row_threshold:
        if has_finsum_marker and len(paths) <= _FINSUM_MAX_FILES:
            return "fin_report_summary", code_hits, con_hits
        return "fin_report", code_hits, con_hits
    if con_hits > 0:
        return "con_call", code_hits, con_hits
    return None, code_hits, con_hits


_SHEET_NAME_INVALID_RE = re.compile(r"[:\\/?*\[\]]")


def sheet_name(folder, taken):
    """Excel worksheet names: max 31 chars, no : \\ / ? * [ ]. Derived from
    the folder's own basename so each sheet is recognizable without opening
    it; de-duplicated against `taken` in case two folders' basenames collide
    after truncation/sanitizing."""
    name = _SHEET_NAME_INVALID_RE.sub("_", Path(folder).name)[:31] or "sheet"
    base, n = name, 2
    while name in taken:
        suffix = f"_{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    taken.add(name)
    return name


_EXCEL_THOUSANDS_INT_FORMAT = "#,##0"
_EXCEL_THOUSANDS_DECIMAL_FORMAT = "#,##0.00"
# Fin-report monetary rows are divided by 1000 before being written; con-call
# rows are untouched. Three decimals so the division can't silently round
# sub-thousand precision away.
_EXCEL_SCALED_FORMAT = "#,##0.000"
# Excel's native percent type multiplies the stored value by 100 for display,
# but every ratio in this codebase is ALREADY stored as a percent number (65.4
# meaning 65.4%). So the /100 happens ONLY here, at the point of writing an
# Excel percent cell, and nowhere else.
#   → docs/knowledge/cli-and-export.md#excel-的百分比與千分之一
_EXCEL_PERCENT_FORMAT = "0.00%"


def write_excel_merged(sheets, out_path):
    """sheets: list of (sheet_title, rows) where rows are (term, value,
    term_found, page, note, is_percent, is_scaled) tuples. One workbook, one
    sheet per folder.

    The note column carries every self-check's output - burying those in
    console output alone defeats the point of computing them.

    Numeric cells get a real Excel NUMBER FORMAT, never a pre-baked display
    string, so the value stays usable in formulas and sorting."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title=title)
        ws.append(["term", "value", "term_found", "page", "note"])
        for term, value, term_found, page, note, is_percent, is_scaled in rows:
            excel_value = value / 100 if (is_percent and isinstance(value, (int, float))) else value
            ws.append([term, excel_value, term_found, page, note])
            if isinstance(value, (int, float)):
                cell = ws.cell(row=ws.max_row, column=2)
                if is_percent:
                    cell.number_format = _EXCEL_PERCENT_FORMAT
                elif is_scaled:
                    cell.number_format = _EXCEL_SCALED_FORMAT
                elif value == int(value):
                    cell.number_format = _EXCEL_THOUSANDS_INT_FORMAT
                else:
                    cell.number_format = _EXCEL_THOUSANDS_DECIMAL_FORMAT
    wb.save(out_path)
    return out_path


def open_file(path):
    """Open a file with its OS-default application (Windows: os.startfile)."""
    import os
    os.startfile(path)  # noqa: only reached on win32, per this project's environment


def lookup_concall_roa_roe(concall_folder, config_path, bank, verbose):
    """Best-match ROA/ROE from an earnings-call deck, for collect_summary_rows'
    concall fallback. Lives HERE because only this module imports both
    packages - financialReports importing earningsCalls would be a cycle.
      → docs/knowledge/cli-and-export.md#財報與法說會怎麼配對"""
    terms = ec.load_terms(config_path)
    primary_aliases = ec.PRIMARY_BANK_ENTITIES.get(bank) if bank else None
    roa = ec.find_term_value(concall_folder, terms["ROA(稅後年化)"], verbose=verbose,
                              prefer_quarterly=True, primary_aliases=primary_aliases)
    roe = ec.find_term_value(concall_folder, terms["ROE(稅後年化)"], verbose=verbose,
                              prefer_quarterly=True, primary_aliases=primary_aliases)
    return (roa[1] if roa else None), (roe[1] if roe else None)


def fin_report_rows(folder, verbose, concall_folder=None, config_path=_DEFAULT_CONFIG, finsum=False):
    """Returns (kind, rows_for_print_or_None, excel_rows) - kind is 'ok',
    'no_bank' or 'no_layout'. Runs the bank-scoped summary once; callers
    decide whether to print it or fold it into a merged export.

    concall_folder: this bank's paired deck, used ONLY as ROA/ROE's fallback
    source. finsum: routes to collect_summary_rows_finsum, identical except
    for the 稅後淨利 code table."""
    bank = fin.detect_bank(folder)
    if bank is None:
        return "no_bank", None, None
    if verbose:
        print(f"  Auto-detected bank: {bank}")
    concall_roa = concall_roe = None
    if concall_folder is not None:
        concall_roa, concall_roe = lookup_concall_roa_roe(concall_folder, config_path, bank, verbose)
    collect_fn = fin.collect_summary_rows_finsum if finsum else fin.collect_summary_rows
    try:
        rows = collect_fn(folder, bank, verbose=verbose,
                           concall_roa=concall_roa, concall_roe=concall_roe)
    except ValueError:
        # No layout for this filing's industry - skip, exactly as for an
        # unresolvable bank. The alternative is a sheet whose standardized
        # terms don't describe the numbers under them.
        #   → docs/knowledge/industry-and-layout.md#summary_layout-為什麼綁死產業
        return "no_layout", None, None
    # Warned here, NOT at a print site: the csv and excel exits never print
    # the rows, and a batch run is where a wholesale failed read would
    # otherwise pass unnoticed.
    warning = fin.summary_coverage_warning(rows, folder=folder)
    if warning:
        print(f"  {warning}")
    excel_rows = []
    for r in rows:
        is_percent = r.get("is_percent", False)
        value = r["value"]
        # Source documents report NT$ thousands. The /1000 happens HERE, at
        # the export boundary only, so the internal value stays in the
        # documents' own units everywhere else.
        #   → docs/knowledge/cli-and-export.md#excel-的百分比與千分之一
        if not is_percent and isinstance(value, (int, float)):
            value = value / 1000
        excel_rows.append((r["term"], value, r.get("matched_label") or "",
                            fin.page_num(r["source_file"]), r.get("note") or "",
                            is_percent, not is_percent))
    return "ok", rows, excel_rows


def con_call_rows(folder, config_path, verbose):
    terms = ec.load_terms(config_path)
    rows = ec.collect_con_call_summary(folder, terms, verbose=verbose)
    excel_rows = []
    for r in rows:
        found = r.get("matched_label") or ""
        page = fin.page_num(r["source_file"])
        note = r.get("note") or ""
        if r["kind"] == "ratio":
            excel_rows.append((r["term"], r["individual"], found, page, note, True, False))
        else:
            excel_rows.append((r["term"], r["value"], found, page, note, r.get("is_percent", False), False))
    return rows, excel_rows


def run_fin_report(folder, export, verbose, concall_folder=None, config_path=_DEFAULT_CONFIG, finsum=False):
    kind, rows, excel_rows = fin_report_rows(folder, verbose, concall_folder=concall_folder,
                                              config_path=config_path, finsum=finsum)
    if kind == "no_bank":
        # Name the candidates when there ARE some: at batch scale "couldn't
        # detect" alone doesn't say whether this entity is unsupported or
        # merely ambiguous, and only the second is fixable with --bank.
        candidates = fin.bank_candidates(folder)
        reason = (f"several banks are named ({', '.join(candidates)})" if candidates
                  else "no known bank name found")
        print(f"  Couldn't auto-detect the bank for {folder} - {reason} - skipping "
              f"(re-run as `cli.py acct <folder> summary --bank ...` to override).")
        return None
    if kind == "no_layout":
        print(f"  Skipping {folder} - {fin.summary_layout_error(fin.detect_industry_category(folder))} "
              f"(re-run as `cli.py acct <folder> summary --industry ...` to override).")
        return None
    if export == "csv":
        out_path = fin.write_summary_csv(folder, rows)
        print(f"  Wrote {len(rows)} row(s) to {out_path}")
        return None
    if export == "excel":
        return excel_rows
    fin.print_summary_rows(rows)
    return None


def run_con_call(folder, config_path, export, verbose):
    rows, excel_rows = con_call_rows(folder, config_path, verbose)
    if export == "csv":
        out_path = ec.write_summary_csv(folder, rows)
        print(f"  Wrote {len(rows)} row(s) to {out_path}")
        return None
    if export == "excel":
        return excel_rows
    ec.print_summary_rows(rows)
    return None


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("folders", nargs="*",
                     help="Up to 2 folders to classify. If omitted, the folder-picker dialog opens "
                          "(up to twice - cancel the second one if you only have one folder).")
    ap.add_argument("--config", default=_DEFAULT_CONFIG,
                     help="Path to the con-call term config JSON (default: bundled con_call_terms.json)")
    ap.add_argument("--export", choices=["csv", "excel"],
                     help="Write each folder's results to a file (term | value | term_found | page "
                          "columns for excel) instead of stdout")
    ap.add_argument("-v", "--verbose", action="store_true", help="Print per-file/per-term/per-code detail")
    args = ap.parse_args()

    folders = args.folders if args.folders else pick_folders(max_count=2)
    # Same folder picked twice (easy to do in the dialog) would otherwise run
    # and export the same extraction twice; keep first-seen order.
    seen, deduped = set(), []
    for f in folders:
        key = str(Path(f).resolve()).lower()
        if key not in seen:
            seen.add(key)
            deduped.append(f)
    folders = deduped
    if not folders:
        ap.error("No folder selected.")
    if len(folders) > 2:
        ap.error("At most 2 folders are supported.")

    codes = load_all_codes()

    # Classify every folder FIRST, before running anything, so a fin_report
    # folder can be paired with its con_call sibling. Only exactly-one-of-each
    # is an unambiguous pairing; anything else falls through.
    #   → docs/knowledge/cli-and-export.md#財報與法說會怎麼配對
    classified = []  # [(folder, kind, code_hits, con_hits), ...]
    for folder in folders:
        kind, code_hits, con_hits = classify_folder(folder, codes)
        classified.append((folder, kind, code_hits, con_hits))

    fin_folders = [f for f, k, _, _ in classified if k in ("fin_report", "fin_report_summary")]
    con_folders = [f for f, k, _, _ in classified if k == "con_call"]
    paired_con_folder = con_folders[0] if len(fin_folders) == 1 and len(con_folders) == 1 else None
    paired_fin_folder = fin_folders[0] if len(fin_folders) == 1 and len(con_folders) == 1 else None

    taken_sheet_names = set()
    excel_sheets = []  # [(sheet_title, rows), ...] - one merged workbook, not per-folder files
    pending_fin_rows = pending_con_rows = None  # held back when paired, to merge into one sheet below

    for folder, kind, code_hits, con_hits in classified:
        print(f"\n=== {folder} ===")
        if kind is None:
            print(f"  Couldn't classify this folder (coded statement rows: {code_hits}, "
                  f"con-call cover markers: {con_hits}) - skipping. Run `cli.py acct` or "
                  f"`cli.py call` on it instead.")
            continue
        if kind == "fin_report":
            print(f"  Detected: financial report ({code_hits} coded statement row(s) found)")
            excel_rows = run_fin_report(folder, args.export, args.verbose,
                                         concall_folder=paired_con_folder, config_path=args.config)
        elif kind == "fin_report_summary":
            print(f"  Detected: summarized fin report ({code_hits} coded statement row(s) found, "
                  f"'{_FINSUM_MARKER}' marker present)")
            excel_rows = run_fin_report(folder, args.export, args.verbose,
                                         concall_folder=paired_con_folder, config_path=args.config,
                                         finsum=True)
        else:
            print(f"  Detected: earnings-call deck ({con_hits} file(s) with con-call markers)")
            excel_rows = run_con_call(folder, args.config, args.export, args.verbose)
        if excel_rows is None:
            continue
        if args.export == "excel" and paired_fin_folder is not None and folder == paired_fin_folder:
            pending_fin_rows = excel_rows
        elif args.export == "excel" and paired_con_folder is not None and folder == paired_con_folder:
            pending_con_rows = excel_rows
        else:
            excel_sheets.append((sheet_name(folder, taken_sheet_names), excel_rows))

    if pending_fin_rows is not None and pending_con_rows is not None:
        merged_rows = merge_fin_and_con_rows(pending_fin_rows, pending_con_rows)
        excel_sheets.append((sheet_name(paired_fin_folder, taken_sheet_names), merged_rows))
    else:
        # One side produced no rows. Whatever WAS held back for the merge
        # still has to reach the workbook as its own sheet - without this the
        # user gets an Excel file missing an entire folder, silently.
        for folder, rows in ((paired_fin_folder, pending_fin_rows),
                             (paired_con_folder, pending_con_rows)):
            if rows is not None:
                excel_sheets.append((sheet_name(folder, taken_sheet_names), rows))

    if args.export == "excel" and excel_sheets:
        # One merged file to Downloads, then opened immediately.
        #   → docs/knowledge/cli-and-export.md#匯出檔為什麼在-downloads
        out_path = _EXCEL_EXPORT_DIR / "runfinder_export.xlsx"
        try:
            write_excel_merged(excel_sheets, out_path)
        except PermissionError:
            # The previous export is still open in Excel and locked - fall
            # back to a timestamped name rather than crashing the run.
            import datetime
            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = _EXCEL_EXPORT_DIR / f"runfinder_export_{stamp}.xlsx"
            write_excel_merged(excel_sheets, out_path)
        print(f"\nWrote {len(excel_sheets)} sheet(s) to {out_path}")
        open_file(out_path)


# The single user-facing entry point. Dispatched by PEEKING AT ARGV, not with
# argparse subparsers: cli's own CLI takes a bare positional, and wrapping
# subparsers around that changes the interface people already use.
#   → docs/knowledge/cli-and-export.md#子命令為什麼用-peek-argv-而不是-subparsers
_SUBCOMMANDS = {
    "acct": ("financialReports.statements", "per-statement / summary extraction from a filing"),
    "call": ("earningsCalls.summary", "term extraction from an earnings-call deck"),
    "npl": ("regulatorDatasets.disclosures", "fetch the FSC 銀行局 monthly datasets"),
}


def _dispatch(argv):
    """Hand off to a package's own CLI, or return False to run cli's."""
    if len(argv) < 2 or argv[1] not in _SUBCOMMANDS:
        return False
    module_name, _help = _SUBCOMMANDS[argv.pop(1)]
    # argv is sys.argv - the subcommand is removed above so the target's own
    # argparse sees exactly the arguments it always did.
    importlib.import_module(module_name).main()
    return True


if __name__ == "__main__":
    if not _dispatch(sys.argv):
        main()
