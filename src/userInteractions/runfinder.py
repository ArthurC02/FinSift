"""
Pick up to 2 folders of converted markdown (.md) files via a folder-picker
dialog and auto-detect which one is a financial report (balance sheet /
income statement / cash flow) and which one is an earnings-call deck (法說
會 / 法人說明會), then run the right extractor's curated summary on each -
so you don't have to remember which of acctfinder.py / callfinder.py to run
or pass folders by hand.

Detection:
  - financial-report signal: table rows whose first cell is an exact code
    from ANY of the 3 industry coding files (金控業/金融業/保險業.xlsx -
    e.g. "10000", "41000") - authoritative, since only genuine coded
    financial statements have these; con-call decks sometimes use the same
    section NAMES (e.g. "資產負債表") as a plain row LABEL in a summary
    table, which would false-positive on a text-only marker search.
  - con-call signal: 法說會, 法人說明會, 說明會, 投資人簡報, 法人電話會議
    appearing in any file (typically the cover/title page).
  Coded rows win if both signals are present, since they're structural
  evidence rather than a text mention.

Usage:
    python runfinder.py [--config con_call_terms.json] [--export csv] [-v]

With no arguments, opens the folder-picker up to twice (cancel the second
dialog if you only have one folder to classify).
"""

import argparse
import importlib
import re
import sys
from pathlib import Path

# `python src/userInteractions/runfinder.py` puts THIS directory on sys.path,
# not src/, so the sibling packages would not resolve. This is the one file
# users are told to run by path (it is the single CLI - see _SUBCOMMANDS), so
# it bootstraps src/ itself rather than making everyone set PYTHONPATH.
_SRC = Path(__file__).resolve().parent.parent
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

from financialReports import acctfinder as af
from earningsCalls import callfinder as cf

# Windows consoles often default to cp1252, which can't print CJK output.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

# Repo root is THREE levels up from src/<package>/, not two. This module
# moved down a directory; the same expression silently pointed at src/
# instead. See core/industry.py, where exactly this bit once.
_DEFAULT_CONFIG = str(Path(__file__).resolve().parent.parent.parent / "data" / "con_call_terms.json")
_EXCEL_EXPORT_DIR = Path.home() / "Downloads"

_CON_CALL_MARKERS = ["法說會", "法人說明會", "說明會", "說明簡報", "投資人簡報",
                     "投資人關係", "法人電話會議"]


def load_all_codes():
    """Union of every code across all 3 industry coding files and all 3
    statements - used only to recognize "this table row's leading cell is
    SOME real account code" for classify_folder(), not to resolve a
    specific industry's dictionary (that happens later, per-folder, once
    a folder is known to be a financial report - see fin_report_rows)."""
    codes = set()
    for path in af.INDUSTRY_CODING_FILES.values():
        for stmt in af.STATEMENTS:
            codes.update(af.load_code_dictionary(path, stmt).keys())
    return codes


def pick_folders(max_count=2):
    """Open the folder-picker dialog up to max_count times, stopping early
    if the user cancels (so picking just 1 folder is normal, not an error).
    Opens in Downloads by default - Windows' native dialog otherwise
    remembers whatever directory some OTHER, unrelated program last used it
    in (e.g. an old project's folder) rather than where these .md
    conversions actually live."""
    import tkinter as tk
    from tkinter import filedialog
    folders = []
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    downloads = Path.home() / "Downloads"
    initialdir = str(downloads) if downloads.is_dir() else None
    for i in range(max_count):
        title = f"Select folder {i + 1} of up to {max_count} (Cancel if done)"
        folder = filedialog.askdirectory(title=title, initialdir=initialdir)
        if not folder:
            break
        folders.append(folder)
    root.destroy()
    return folders


# Distinctive of the quarterly SUMMARIZED fin-report website disclosure
# (依「公開發行銀行財務報告編製準則」第三十二條) - this exact ratio line isn't
# part of any of the 3 core statements a FULL individual filing's 150-280
# files revolve around, confirmed present in all of CTBC/北富銀/國泰's real
# 114Q4 summarized disclosures. Combined with a low file-count ceiling (a
# full filing is 150-280 files; the summarized one is ~10-15) so a real full
# filing that happens to also disclose this ratio somewhere within its many
# files still classifies as 'fin_report', not 'fin_report_summary'.
_FINSUM_MARKER = "活期性存款比率"
_FINSUM_MAX_FILES = 30

# When a fin_report/fin_report_summary folder is paired 1:1 with a con_call
# folder (see paired_con_folder/paired_fin_folder in main()), their rows are
# combined into ONE sheet instead of two, in this fixed business order:
# every fin_report row first (acctfinder.SUMMARY_LAYOUT's own order, ending
# 活存比/CIR pulled after ROA/ROE per request), then every con_call row
# (callfinder.RATIO_TERMS/BALANCE_TERMS/NPL_RATIO_TERM/NPL_COVERAGE_TERM).
# Deliberately spelled out in full (not "fin terms then leftover con terms")
# so this list stays the single source of truth for the merged sheet's row
# order - merge_fin_and_con_rows falls back to source order only for a term
# that isn't listed here at all (future SUMMARY_LAYOUT/con-call additions).
_MERGED_TERM_ORDER = [
    "總資產", "淨收益", "利息淨收益", "手續費淨收益", "評價及已實現",
    "其他非利息收益", "營業費用", "員工福利費用", "折舊及攤銷費用",
    "其他費用", "呆帳提存(收回)", "稅前淨利", "所得稅費用", "稅後淨利",
    "ROA", "ROE", "活存比", "CIR",
    "NIM", "存放利差", "放款均率", "存款均率",
    "逾放比率", "備抵呆帳/逾期放款",
    "企業放款", "房貸", "個人放款", "信用卡循環",
    "法說會放款餘額合計", "法說會外幣放款",
]


def merge_fin_and_con_rows(fin_rows, con_rows):
    """fin_rows/con_rows: excel_rows lists (see fin_report_rows/con_call_rows)
    from one paired fin_report(+summary) folder and one con_call folder.
    Returns a single reordered list per _MERGED_TERM_ORDER, each term
    appearing exactly once even though it could in principle exist on both
    sides (it never does in practice, but fin_rows is checked first since
    活存比/CIR only live there)."""
    fin_by_term = {r[0]: r for r in fin_rows}
    con_by_term = {r[0]: r for r in con_rows}
    used_fin, used_con = set(), set()
    merged = []
    for term in _MERGED_TERM_ORDER:
        if term in fin_by_term:
            merged.append(fin_by_term[term])
            used_fin.add(term)
            # A con-call row for the same term is superseded, not still
            # pending: without this it fell through to the trailing extend
            # below and the term was emitted a second time, contradicting
            # this function's own "each term appearing exactly once".
            used_con.add(term)
        elif term in con_by_term:
            merged.append(con_by_term[term])
            used_con.add(term)
    merged.extend(r for r in fin_rows if r[0] not in used_fin)
    merged.extend(r for r in con_rows if r[0] not in used_con)
    return merged


def classify_folder(folder, codes, code_row_threshold=5):
    """Return ('fin_report' | 'fin_report_summary' | 'con_call' | None,
    code_hits, con_hits). code_hits counts table rows whose first cell is
    an exact account code; con_hits counts files containing a con-call
    cover-page marker. 'fin_report_summary' (see _FINSUM_MARKER) routes to
    acctfinder.collect_summary_rows_finsum() instead of the full-filing
    collect_summary_rows() - same SUMMARY_LAYOUT items, different per-bank
    code scheme for 稅後淨利 (see SUMMARY_CODE_OVERRIDES_FINSUM)."""
    code_hits = 0
    con_hits = 0
    has_finsum_marker = False
    # rglob, matching detect_bank / find_code_value / collect_statement_rows /
    # find_term_value. With glob, a folder whose .md files sat in a subdirectory
    # classified as None and was skipped, while acctfinder run directly on the
    # same folder worked fine.
    paths = sorted(Path(folder).rglob("*.md"))
    for path in paths:
        text = path.read_text(encoding="utf-8", errors="ignore")
        if any(m in text for m in _CON_CALL_MARKERS):
            con_hits += 1
        if _FINSUM_MARKER in text:
            has_finsum_marker = True
        for line in text.splitlines():
            if not line.startswith("|"):
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if cells and cells[0] in codes:
                code_hits += 1
    if code_hits >= code_row_threshold:
        if has_finsum_marker and len(paths) <= _FINSUM_MAX_FILES:
            return "fin_report_summary", code_hits, con_hits
        return "fin_report", code_hits, con_hits
    if con_hits > 0:
        return "con_call", code_hits, con_hits
    return None, code_hits, con_hits


_SHEET_NAME_INVALID_RE = re.compile(r"[:\\/?*\[\]]")


def sheet_name(folder, taken):
    """Excel worksheet names: max 31 chars, no : \\ / ? * [ ]. Derived from
    the folder's own basename so each sheet is recognizable without opening
    it; de-duplicated against `taken` in case two folders' basenames collide
    after truncation/sanitizing."""
    name = _SHEET_NAME_INVALID_RE.sub("_", Path(folder).name)[:31] or "sheet"
    base, n = name, 2
    while name in taken:
        suffix = f"_{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    taken.add(name)
    return name


_EXCEL_THOUSANDS_INT_FORMAT = "#,##0"
_EXCEL_THOUSANDS_DECIMAL_FORMAT = "#,##0.00"
# Fin-report (full and summarized) monetary rows are divided by 1000 before
# being written (see fin_report_rows' _scale_fin_report_rows) - con-call
# rows are untouched. Always shown to 3 decimals so sub-thousand precision
# isn't silently rounded away by the division.
_EXCEL_SCALED_FORMAT = "#,##0.000"
# A real Excel percentage type multiplies the stored number by 100 for
# display (its underlying value is a fraction, e.g. 0.654 for "65.40%") -
# but every ratio value in this codebase is already stored AS a percent
# number (65.4 meaning 65.4%, per format_pct() throughout acctfinder.py/
# callfinder.py). So the value is divided by 100 ONLY here, at the point of
# writing an Excel-native percent cell, and nowhere else - the internal
# "already-times-100" convention everywhere else is untouched.
_EXCEL_PERCENT_FORMAT = "0.00%"


def write_excel_merged(sheets, out_path):
    """sheets: list of (sheet_title, rows) where rows are (term, value,
    term_found, page, note, is_percent, is_scaled) tuples. Each folder gets its own
    sheet in ONE workbook, rather than a separate file per folder. The note
    column carries the self-check output (balance-sheet identity, ROA/ROE
    cross-check and plausibility bounds, loan-book reconciliation, and the
    "regulator month not published yet" fallback flag) - burying those in
    console output only would defeat the point of computing them.
    Numeric cells get a real Excel number format applied (not a pre-baked
    display string) so the value stays usable in formulas/sorting: ratio
    rows (is_percent) as a genuine Excel percentage (see
    _EXCEL_PERCENT_FORMAT's docs on the /100 needed to make that correct),
    everything else with thousands-separator grouping."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title=title)
        ws.append(["term", "value", "term_found", "page", "note"])
        for term, value, term_found, page, note, is_percent, is_scaled in rows:
            excel_value = value / 100 if (is_percent and isinstance(value, (int, float))) else value
            ws.append([term, excel_value, term_found, page, note])
            if isinstance(value, (int, float)):
                cell = ws.cell(row=ws.max_row, column=2)
                if is_percent:
                    cell.number_format = _EXCEL_PERCENT_FORMAT
                elif is_scaled:
                    cell.number_format = _EXCEL_SCALED_FORMAT
                elif value == int(value):
                    cell.number_format = _EXCEL_THOUSANDS_INT_FORMAT
                else:
                    cell.number_format = _EXCEL_THOUSANDS_DECIMAL_FORMAT
    wb.save(out_path)
    return out_path


def open_file(path):
    """Open a file with its OS-default application (Windows: os.startfile)."""
    import os
    os.startfile(path)  # noqa: only reached on win32, per this project's environment


def lookup_concall_roa_roe(concall_folder, config_path, bank, verbose):
    """Best-match ROA(稅後年化)/ROE(稅後年化) from an earnings-call deck, for
    collect_summary_rows' concall fallback - looked up here (not inside
    acctfinder.py) since only this module imports both acctfinder.py and
    callfinder.py; acctfinder.py importing callfinder.py would be circular
    (callfinder.py already imports FROM acctfinder.py)."""
    terms = cf.load_terms(config_path)
    primary_aliases = cf.PRIMARY_BANK_ENTITIES.get(bank) if bank else None
    roa = cf.find_term_value(concall_folder, terms["ROA(稅後年化)"], verbose=verbose,
                              prefer_quarterly=True, primary_aliases=primary_aliases)
    roe = cf.find_term_value(concall_folder, terms["ROE(稅後年化)"], verbose=verbose,
                              prefer_quarterly=True, primary_aliases=primary_aliases)
    return (roa[1] if roa else None), (roe[1] if roe else None)


def fin_report_rows(folder, verbose, concall_folder=None, config_path=_DEFAULT_CONFIG, finsum=False):
    """Returns (kind, rows_for_print_or_None, excel_rows) - kind is 'ok' or
    'no_bank'. Runs the bank-scoped acctfinder summary once; callers decide
    whether to print it or fold it into a merged export. 'summary' mode
    never touches an industry coding dictionary for its main SUMMARY_LAYOUT
    rows (it matches raw document codes directly) - only the trailing ROA/
    ROE rows' manual-formula cross-check needs one, auto-detected inside
    collect_summary_rows. Its displayed term is now a fixed/standardized
    name (SUMMARY_LAYOUT), not the document's own wording - that's kept
    separately as matched_label. concall_folder: this bank's paired
    earnings-call deck, if runfinder classified one alongside this fin
    report folder - used only as ROA/ROE's fallback source (see
    collect_summary_rows/collect_roa_roe's priority order). finsum: True
    when classify_folder identified this as the quarterly SUMMARIZED
    disclosure (see _FINSUM_MARKER) - routes to
    collect_summary_rows_finsum() instead, which is otherwise identical
    except for a different 稅後淨利 code table."""
    bank = af.detect_bank(folder)
    if bank is None:
        return "no_bank", None, None
    if verbose:
        print(f"  Auto-detected bank: {bank}")
    concall_roa = concall_roe = None
    if concall_folder is not None:
        concall_roa, concall_roe = lookup_concall_roa_roe(concall_folder, config_path, bank, verbose)
    collect_fn = af.collect_summary_rows_finsum if finsum else af.collect_summary_rows
    try:
        rows = collect_fn(folder, bank, verbose=verbose,
                           concall_roa=concall_roa, concall_roe=concall_roe)
    except ValueError:
        # summary mode has no layout for this filing's industry. Skipping is
        # the same handling a folder gets when its bank can't be resolved -
        # the alternative is emitting a sheet whose standardized terms don't
        # describe the numbers under them. The message is rebuilt at the
        # print site rather than carried in the rows slot, matching how the
        # no_bank branch re-derives its candidates.
        return "no_layout", None, None
    # Warned here, not at a print site: run_fin_report's csv and excel exits
    # never print the rows, and a batch run is exactly where a wholesale
    # failed read would otherwise pass unnoticed.
    warning = af.summary_coverage_warning(rows, folder=folder)
    if warning:
        print(f"  {warning}")
    # r["term"] is the canonical/standardized display name (SUMMARY_LAYOUT),
    # r["matched_label"] is the document's own wording for that row (or
    # None for an N/A row, or the term itself for a composite - see
    # collect_summary_rows).
    excel_rows = []
    for r in rows:
        is_percent = r.get("is_percent", False)
        value = r["value"]
        # Balance/income-statement amounts are reported in NT$ thousands by
        # every source document; divide by 1000 (-> NT$ millions) here, at
        # the export boundary, so the internal value stays in the units the
        # documents themselves use everywhere else (self-checks, printed
        # summaries, CSV export). Ratios (is_percent) are never amounts, so
        # they're left untouched.
        if not is_percent and isinstance(value, (int, float)):
            value = value / 1000
        excel_rows.append((r["term"], value, r.get("matched_label") or "",
                            af.page_num(r["source_file"]), r.get("note") or "",
                            is_percent, not is_percent))
    return "ok", rows, excel_rows


def con_call_rows(folder, config_path, verbose):
    terms = cf.load_terms(config_path)
    rows = cf.collect_con_call_summary(folder, terms, verbose=verbose)
    excel_rows = []
    for r in rows:
        found = r.get("matched_label") or ""
        page = af.page_num(r["source_file"])
        note = r.get("note") or ""
        if r["kind"] == "ratio":
            excel_rows.append((r["term"], r["individual"], found, page, note, True, False))
        else:
            excel_rows.append((r["term"], r["value"], found, page, note, r.get("is_percent", False), False))
    return rows, excel_rows


def run_fin_report(folder, export, verbose, concall_folder=None, config_path=_DEFAULT_CONFIG, finsum=False):
    kind, rows, excel_rows = fin_report_rows(folder, verbose, concall_folder=concall_folder,
                                              config_path=config_path, finsum=finsum)
    if kind == "no_bank":
        # Name the candidates when there ARE some: at batch scale "couldn't
        # detect" alone doesn't say whether this entity is unsupported or
        # merely ambiguous, and only the second is fixable with --bank.
        candidates = af.bank_candidates(folder)
        reason = (f"several banks are named ({', '.join(candidates)})" if candidates
                  else "no known bank name found")
        print(f"  Couldn't auto-detect the bank for {folder} - {reason} - skipping "
              f"(run acctfinder.py directly with --bank to override).")
        return None
    if kind == "no_layout":
        print(f"  Skipping {folder} - {af.summary_layout_error(af.detect_industry_category(folder))} "
              f"(run acctfinder.py directly with --industry to override).")
        return None
    if export == "csv":
        out_path = af.write_summary_csv(folder, rows)
        print(f"  Wrote {len(rows)} row(s) to {out_path}")
        return None
    if export == "excel":
        return excel_rows
    af.print_summary_rows(rows)
    return None


def run_con_call(folder, config_path, export, verbose):
    rows, excel_rows = con_call_rows(folder, config_path, verbose)
    if export == "csv":
        out_path = cf.write_summary_csv(folder, rows)
        print(f"  Wrote {len(rows)} row(s) to {out_path}")
        return None
    if export == "excel":
        return excel_rows
    cf.print_summary_rows(rows)
    return None


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("folders", nargs="*",
                     help="Up to 2 folders to classify. If omitted, the folder-picker dialog opens "
                          "(up to twice - cancel the second one if you only have one folder).")
    ap.add_argument("--config", default=_DEFAULT_CONFIG,
                     help="Path to the con-call term config JSON (default: bundled con_call_terms.json)")
    ap.add_argument("--export", choices=["csv", "excel"],
                     help="Write each folder's results to a file (term | value | term_found | page "
                          "columns for excel) instead of stdout")
    ap.add_argument("-v", "--verbose", action="store_true", help="Print per-file/per-term/per-code detail")
    args = ap.parse_args()

    folders = args.folders if args.folders else pick_folders(max_count=2)
    # Same folder picked twice (easy to do in the dialog) would otherwise run
    # and export the same extraction twice; keep first-seen order.
    seen, deduped = set(), []
    for f in folders:
        key = str(Path(f).resolve()).lower()
        if key not in seen:
            seen.add(key)
            deduped.append(f)
    folders = deduped
    if not folders:
        ap.error("No folder selected.")
    if len(folders) > 2:
        ap.error("At most 2 folders are supported.")

    codes = load_all_codes()

    # Classify every folder FIRST (before running anything) so a fin_report
    # folder can be paired with its own con_call sibling, if both were
    # selected together - needed only for ROA/ROE's concall fallback (see
    # fin_report_rows/collect_roa_roe); with exactly one of each, the
    # pairing is unambiguous. With 0 or 2+ of either kind, there's nothing
    # sound to pair, so ROA/ROE just falls through to its other sources.
    classified = []  # [(folder, kind, code_hits, con_hits), ...]
    for folder in folders:
        kind, code_hits, con_hits = classify_folder(folder, codes)
        classified.append((folder, kind, code_hits, con_hits))

    fin_folders = [f for f, k, _, _ in classified if k in ("fin_report", "fin_report_summary")]
    con_folders = [f for f, k, _, _ in classified if k == "con_call"]
    paired_con_folder = con_folders[0] if len(fin_folders) == 1 and len(con_folders) == 1 else None
    paired_fin_folder = fin_folders[0] if len(fin_folders) == 1 and len(con_folders) == 1 else None

    taken_sheet_names = set()
    excel_sheets = []  # [(sheet_title, rows), ...] - one merged workbook, not per-folder files
    pending_fin_rows = pending_con_rows = None  # held back when paired, to merge into one sheet below

    for folder, kind, code_hits, con_hits in classified:
        print(f"\n=== {folder} ===")
        if kind is None:
            print(f"  Couldn't classify this folder (coded statement rows: {code_hits}, "
                  f"con-call cover markers: {con_hits}) - skipping. Run acctfinder.py or "
                  f"callfinder.py directly on it instead.")
            continue
        if kind == "fin_report":
            print(f"  Detected: financial report ({code_hits} coded statement row(s) found)")
            excel_rows = run_fin_report(folder, args.export, args.verbose,
                                         concall_folder=paired_con_folder, config_path=args.config)
        elif kind == "fin_report_summary":
            print(f"  Detected: summarized fin report ({code_hits} coded statement row(s) found, "
                  f"'{_FINSUM_MARKER}' marker present)")
            excel_rows = run_fin_report(folder, args.export, args.verbose,
                                         concall_folder=paired_con_folder, config_path=args.config,
                                         finsum=True)
        else:
            print(f"  Detected: earnings-call deck ({con_hits} file(s) with con-call markers)")
            excel_rows = run_con_call(folder, args.config, args.export, args.verbose)
        if excel_rows is None:
            continue
        if args.export == "excel" and paired_fin_folder is not None and folder == paired_fin_folder:
            pending_fin_rows = excel_rows
        elif args.export == "excel" and paired_con_folder is not None and folder == paired_con_folder:
            pending_con_rows = excel_rows
        else:
            excel_sheets.append((sheet_name(folder, taken_sheet_names), excel_rows))

    if pending_fin_rows is not None and pending_con_rows is not None:
        merged_rows = merge_fin_and_con_rows(pending_fin_rows, pending_con_rows)
        excel_sheets.append((sheet_name(paired_fin_folder, taken_sheet_names), merged_rows))
    else:
        # One side produced no rows - e.g. run_fin_report returned None because
        # the bank couldn't be auto-detected. Whatever WAS held back for the
        # merge still has to reach the workbook as its own sheet; without this
        # it was dropped silently, and the user got an Excel file missing an
        # entire folder with nothing in the output saying so.
        for folder, rows in ((paired_fin_folder, pending_fin_rows),
                             (paired_con_folder, pending_con_rows)):
            if rows is not None:
                excel_sheets.append((sheet_name(folder, taken_sheet_names), rows))

    if args.export == "excel" and excel_sheets:
        # One merged file (each folder its own sheet), saved to Downloads -
        # not inside any of the source folders, and not their shared parent
        # either (which may not be Downloads) - then opened immediately so
        # the run's result is visible without hunting for it.
        out_path = _EXCEL_EXPORT_DIR / "runfinder_export.xlsx"
        try:
            write_excel_merged(excel_sheets, out_path)
        except PermissionError:
            # The previous export is still open in Excel (auto-open leaves it
            # locked until closed) - fall back to a timestamped name instead
            # of crashing the whole run.
            import datetime
            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = _EXCEL_EXPORT_DIR / f"runfinder_export_{stamp}.xlsx"
            write_excel_merged(excel_sheets, out_path)
        print(f"\nWrote {len(excel_sheets)} sheet(s) to {out_path}")
        open_file(out_path)


# The single user-facing entry point. The other three packages keep their own
# argparse main() exactly as it was - this only routes to them, so every flag
# they already document keeps working and none of them needed touching.
#
# Dispatched by peeking at argv rather than with argparse subparsers, because
# runfinder's own CLI takes a bare positional (`runfinder <folder> <folder>`)
# and adding subparsers around that would have changed the interface people
# already use. A subcommand name is never a folder name, so the peek is
# unambiguous; anything else falls through to the parser below untouched.
_SUBCOMMANDS = {
    "acct": ("financialReports.acctfinder", "per-statement / summary extraction from a filing"),
    "call": ("earningsCalls.callfinder", "term extraction from an earnings-call deck"),
    "npl": ("regulatorDatasets.npl_finder", "fetch the FSC 銀行局 monthly datasets"),
}


def _dispatch(argv):
    """Hand off to a package's own CLI, or return False to run runfinder's."""
    if len(argv) < 2 or argv[1] not in _SUBCOMMANDS:
        return False
    module_name, _help = _SUBCOMMANDS[argv.pop(1)]
    # argv is sys.argv - the subcommand is removed above so the target's own
    # argparse sees exactly the arguments it always did.
    importlib.import_module(module_name).main()
    return True


if __name__ == "__main__":
    if not _dispatch(sys.argv):
        main()
