"""
Fetch two FSC 銀行局 public monthly disclosure datasets for the tracked banks:

  1. 逾期放款總額 (total overdue/NPL loans) - "本國銀行資產品質評估分析統計表"
     unit: 新臺幣百萬元 (NT$ million)
  2. 循環信用餘額 (credit-card revolving balance, i.e. 信用卡循環) -
     "信用卡重要業務及財務資訊揭露", unit: 新臺幣千元 (NT$ thousand)

(2) exists because not every bank's earnings-call deck discloses 信用卡循環
on its own - 中信's deck folds it into 信用貸款與其他 and never breaks it
out, and 國泰's deck reports only an aggregate 信用卡放款 - yet the con-call
loan recomposition needs it as a standalone figure. This regulator dataset
reports it uniformly for every bank, every month, so it's a reliable
fallback when (and only when) a deck lacks its own number.

Standalone by design (no import from acctfinder/callfinder/runfinder) so it
can be tested and scheduled on its own; callfinder.py calls
fetch_for_quarter() from here for the con-call fallback path.

Page layout (confirmed against real downloads, 2026-07):
  - Both pages are plain server-rendered HTML - every month's link is
    already in the page source, no JavaScript rendering needed.
  - The two sites use DIFFERENT filename conventions, so each gets its own
    period-extraction regex rather than one shared guess:
      NPL:         '...統計表115_05(2).xlsx'  -> ROC year 115, month 5
                   (the '(N)' suffix is an inconsistent revised re-upload)
      credit card: '.../11412_信用卡重要資訊揭露.zip' -> ROC 114, month 12
                   (YYYMM concatenated, and the spreadsheet is inside a ZIP
                   alongside an .ods twin - the page labels it "excel&ods")
  - In both cases "latest" is found by parsing every link's (year, month)
    and taking the max, never by constructing a URL from a naming formula.
  - Both spreadsheets are one flat sheet: a header row, then one row per
    bank. Column positions are located by HEADER TEXT, not a fixed index,
    so a future column reordering doesn't silently return the wrong metric.

Usage:
    python npl_finder.py                    # latest month of both datasets
    python npl_finder.py --year 2025 --quarter 4
"""

import argparse
import io
import re
import ssl
import sys
import unicodedata
import urllib.error
import urllib.request
import zipfile
from pathlib import Path

import openpyxl

# Windows consoles often default to cp1252, which can't print CJK output.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

NPL_PAGE_URL = (
    "https://www.banking.gov.tw/ch/home.jsp?id=590&parentpath=0"
    "&mcustomize=multimessage_view.jsp&dataserno=201202130001&dtable=Disclosure"
)
CREDIT_CARD_PAGE_URL = (
    "https://www.banking.gov.tw/ch/home.jsp?id=591&parentpath=0,590"
    "&mcustomize=multimessage_view.jsp&dataserno=21207"
    "&aplistdn=ou=disclosure,ou=multisite,ou=chinese,ou=ap_root,o=fsc,c=tw&dtable=Disclosure"
)
# Backwards-compatible alias for the original single-dataset name.
# ponytail: zero references inside this repo. Kept because it is explicitly a
# compatibility name and nothing here can see whether an external script
# imports it - delete once that's confirmed.
DISCLOSURE_PAGE_URL = NPL_PAGE_URL

# Exact legal names as they appear in BOTH spreadsheets' own bank-name
# column - confirmed identical across the two datasets.
TARGET_BANKS = ["台北富邦商業銀行", "國泰世華商業銀行", "玉山商業銀行", "中國信託商業銀行"]

NPL_BANK_HEADER = "銀行別"
NPL_VALUE_HEADER = "逾期放款總額"
NPL_UNIT = "新臺幣百萬元"

CC_BANK_HEADER = "金融機構名稱"
CC_VALUE_HEADER = "循環信用餘額"
CC_UNIT = "新臺幣千元"

# 逾放比率/備抵呆帳(逾期放款覆蓋率) live on the SAME "本國銀行資產品質評估
#分析統計表" sheet as 逾期放款總額 (NPL_VALUE_HEADER) - confirmed against a
# real download: these are bank-WIDE ratios (all loans, not credit-card
# specific), so they belong with fetch_overdue_loans(), not the credit-card
# sheet. An earlier version of this module wrongly sourced two
# similarly-named columns from the CREDIT-CARD sheet instead (逾期三個月以上
# 帳款比率/備抵呆帳提足率) - a real, but different, credit-card-only metric
# that doesn't match this sheet's bank-wide 中國信託 values (0.13%/950.93%
# for 114年12月 vs the credit-card sheet's 0.12%/299.62%) - fixed per
# user-supplied ground truth from the actual government page.
# Both already stored as percent-scale numbers in the sheet - never
# divide/multiply these again, just append '%' on display (see format_pct
# in acctfinder.py/callfinder.py).
NPL_RATIO_HEADER = "逾放比率(%)"
NPL_COVERAGE_HEADER = "備抵呆帳/逾期放款(%)"

_CACHE_DIR = Path(__file__).parent.parent / "npl_cache"

_USER_AGENT = "Mozilla/5.0 (compatible; npl_finder/2.0)"

_NPL_PERIOD_RE = re.compile(r"(\d{3})_(\d{1,2})(?:\(\d+\))?\.xlsx", re.IGNORECASE)
# Matched against the RAW (still percent-encoded) href, since the YYYMM_
# prefix sits before the encoded Chinese filename text.
_CC_PERIOD_RE = re.compile(r"/(\d{3})(\d{2})_[^/]*\.zip", re.IGNORECASE)


def _fetch_url(url, timeout=60):
    """GET `url` and return the raw response bytes. Uses the system's
    default (verifying) SSL context - if this raises SSLCertVerificationError
    in your environment, that means your machine's CA trust store is
    missing an intermediate certificate (a local environment issue, common
    on some Windows Python installs - reproduced in testing this module),
    NOT that the connection should be downgraded to unverified. Fix it via
    `pip install pip-system-certs` (uses the Windows store) or certifi -
    don't disable verification, since that reopens this to a MITM attack.
    Raises RuntimeError with that guidance (urlopen otherwise wraps SSL
    errors inside urllib.error.URLError, not ssl.SSLError, so a caller
    catching only ssl.SSLError would miss it - confirmed by testing)."""
    req = urllib.request.Request(url, headers={"User-Agent": _USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read()
    except urllib.error.URLError as e:
        if isinstance(e.reason, ssl.SSLCertVerificationError):
            raise RuntimeError(
                f"SSL certificate verification failed connecting to {url} - this is a local "
                "CA trust store issue (missing intermediate certificate), not a problem with "
                "the site itself. Fix: `pip install pip-system-certs` (or certifi). Do not "
                "disable SSL verification - that removes protection against a "
                f"man-in-the-middle attack. Underlying error: {e.reason}"
            ) from e
        raise RuntimeError(f"Failed to reach {url}: {e}") from e


def _list_period_links(page_url, period_re, ext):
    """Return {(roc_year, month): url} for every dataset file linked on
    `page_url` whose filename matches `period_re`. `ext` narrows the href
    scan to one file type first (e.g. '.xlsx', '.zip') so the NPL page's
    PDF twins - same period, wrong format - can't win the dict slot."""
    html = _fetch_url(page_url).decode("utf-8", errors="ignore")
    hrefs = re.findall(r'href="([^"]+%s)"' % re.escape(ext), html, re.IGNORECASE)
    out = {}
    for href in hrefs:
        m = period_re.search(href)
        if m:
            out[(int(m.group(1)), int(m.group(2)))] = href
    if not out:
        raise RuntimeError(
            f"No dataset links matching {period_re.pattern} found on {page_url} - "
            "the page structure may have changed."
        )
    return out


def resolve_period(links, roc_year=None, month=None):
    """Pick which published month to use. Returns (year, month, url, exact).
    With no requested period, takes the newest available. With one, returns
    it if published; otherwise falls back to the newest month at or BEFORE
    the request (never a later one - a future month's figures are not a
    substitute for the quarter you asked about) and reports exact=False so
    the caller can say which period it actually used. These datasets lag by
    a month or two, so asking for a just-ended quarter routinely lands
    here rather than being an error.

    One case can't honour "never later": a request OLDER than everything
    published. There is nothing at or before it, so the oldest available month
    is returned - which IS later than what was asked for. exact=False and the
    caller's note both disclose it; the docstring used to flatly contradict
    it."""
    if roc_year is None or month is None:
        pick = max(links)
        return pick[0], pick[1], links[pick], True
    key = (roc_year, month)
    if key in links:
        return roc_year, month, links[key], True
    earlier = [k for k in links if k <= key]
    pick = max(earlier) if earlier else min(links)
    return pick[0], pick[1], links[pick], False


def download_file(url, cache_dir=_CACHE_DIR):
    """Download `url` into `cache_dir` (named after the URL's own filename)
    and return its local Path. Skips the download if already cached, so
    repeat runs within a month don't re-fetch; delete the cached file to
    force a refresh if the source re-uploads a revision under the same name."""
    cache_dir = Path(cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)
    dest = cache_dir / urllib.request.unquote(url.rsplit("/", 1)[-1])
    if dest.exists():
        return dest
    dest.write_bytes(_fetch_url(url))
    return dest


def _xlsx_from_zip(zip_path):
    """The credit-card dataset ships as a ZIP holding an .xlsx and an .ods
    twin. Extract the .xlsx next to the archive and return its path."""
    with zipfile.ZipFile(zip_path) as z:
        names = [n for n in z.namelist() if n.lower().endswith(".xlsx")]
        if not names:
            raise RuntimeError(f"No .xlsx inside {zip_path.name} (members: {z.namelist()})")
        dest = zip_path.parent / Path(names[0]).name
        if not dest.exists():
            dest.write_bytes(z.read(names[0]))
        return dest


def _find_header_column(ws, header_text, max_scan_rows=10):
    """1-indexed column of the cell (or short run of stacked cells) whose
    text equals `header_text` once stripped of whitespace. Two passes:
      1. A plain single cell, in any of the first `max_scan_rows` rows,
         exactly equal to `header_text` - the common case (e.g. '銀行別',
         '循環信用餘額').
      2. If nothing matched, a 3-row-STACKED header: some real sheets split
         one metric's name across 3 consecutive rows of the same column
         (confirmed in a real download - the NPL sheet's '逾放'/'比率'/'(%)'
         only read as '逾放比率(%)' once joined). The stacked header band
         is located as the first row where at least 2 columns simultaneously
         hold string text (distinguishing it from single-column title/date/
         unit preamble lines above it, which only populate column 1) -
         concatenating blindly from row 1 would otherwise pull those
         preamble lines into column 1's own "header", never matching
         anything. Only that 3-row band is concatenated per column, never
         reaching down into the data rows below it.
    Not a substring match in either pass, since one metric's name can be a
    substring of another's."""
    target = "".join(header_text.split())
    rows = list(ws.iter_rows(min_row=1, max_row=max_scan_rows))

    for row in rows:
        for cell in row:
            if isinstance(cell.value, str) and "".join(cell.value.split()) == target:
                return cell.column

    header_start = None
    for row in rows:
        if sum(1 for cell in row if isinstance(cell.value, str)) >= 2:
            header_start = row[0].row
            break
    if header_start is None:
        return None
    band = ws.iter_rows(min_row=header_start, max_row=header_start + 2)
    by_col = {}
    for row in band:
        for cell in row:
            if isinstance(cell.value, str):
                by_col.setdefault(cell.column, []).append(cell.value)
    for col, parts in by_col.items():
        if "".join("".join(p.split()) for p in parts) == target:
            return col
    return None


def _parse_number(value):
    """Both sheets are read with data_only=True, but the credit-card one
    stores its figures as TEXT while the NPL one stores real numbers - so
    accept either rather than assuming."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _extract_by_bank(xlsx_path, bank_header, value_header, banks):
    """{bank: value_or_None} for each name in `banks`, locating both the
    bank-name and value columns by header text (see module docstring).
    None means that bank's row wasn't found at all - distinct from a
    present-but-empty cell, which parses to None too but is genuinely
    absent data either way."""
    return {bank: values[value_header]
            for bank, values in _extract_columns_by_bank(xlsx_path, bank_header, [value_header], banks).items()}


def _extract_columns_by_bank(xlsx_path, bank_header, value_headers, banks):
    """{bank: {value_header: value_or_None, ...}} for each name in `banks`,
    reading every header in `value_headers` from the SAME row/workbook load
    (one sheet open instead of one per metric) - used when several metrics
    live in one sheet (e.g. 循環信用餘額 plus the credit-card NPL/coverage
    ratios), so adding a metric never re-downloads or re-parses the file."""
    ws = openpyxl.load_workbook(xlsx_path, data_only=True).worksheets[0]
    bank_col = _find_header_column(ws, bank_header)
    if bank_col is None:
        raise RuntimeError(
            f"Couldn't locate '{bank_header}' header cell in {xlsx_path.name} - "
            "the sheet layout may have changed."
        )
    value_cols = {}
    for header in value_headers:
        col = _find_header_column(ws, header)
        if col is None:
            raise RuntimeError(
                f"Couldn't locate '{header}' header cell in {xlsx_path.name} - "
                "the sheet layout may have changed."
            )
        value_cols[header] = col
    found = {}
    for row in ws.iter_rows(min_row=1):
        name = row[bank_col - 1].value
        if isinstance(name, str) and name.strip() in banks and name.strip() not in found:
            found[name.strip()] = {header: _parse_number(row[col - 1].value)
                                    for header, col in value_cols.items()}
    return {bank: found.get(bank, {h: None for h in value_headers}) for bank in banks}


def _result(kind, year, month, exact, url, unit, values, requested):
    note = ""
    if not exact and requested:
        # Word the reason by direction. A request older than the whole dataset
        # falls back FORWARD to the oldest month available, and calling that
        # "isn't published yet" was simply wrong.
        if (year, month) > tuple(requested):
            why = "that month is no longer in the published dataset"
        else:
            why = "that month isn't published yet"
        note = (f"requested {requested[0]}年{requested[1]}月 but {why} - "
                f"used {year}年{month}月 instead")
    return {"kind": kind, "period": f"{year}年{month}月", "roc_year": year, "month": month,
            "exact": exact, "source_url": url, "unit": unit, "values": values, "note": note}


def fetch_overdue_loans(roc_year=None, month=None, banks=TARGET_BANKS,
                         cache_dir=_CACHE_DIR, verbose=False):
    """逾期放款總額 for `banks`, in 新臺幣百萬元, plus two bank-wide ratios
    read from the SAME sheet (already percent-scale numbers - see
    NPL_RATIO_HEADER/NPL_COVERAGE_HEADER, never rescaled here): 逾放比率
    and 備抵呆帳/逾期放款 (NPL coverage). The result dict's "values" stays
    逾期放款總額-only for backwards compatibility with existing callers;
    "npl_ratios" and "coverage_ratios" carry the two new per-bank dicts
    alongside it."""
    links = _list_period_links(NPL_PAGE_URL, _NPL_PERIOD_RE, ".xlsx")
    year, mon, url, exact = resolve_period(links, roc_year, month)
    if verbose:
        print(f"[逾期放款] using {year}年{mon}月 -> {url}")
    path = download_file(url, cache_dir)
    by_bank = _extract_columns_by_bank(
        path, NPL_BANK_HEADER, [NPL_VALUE_HEADER, NPL_RATIO_HEADER, NPL_COVERAGE_HEADER], banks)
    values = {b: v[NPL_VALUE_HEADER] for b, v in by_bank.items()}
    npl_ratios = {b: v[NPL_RATIO_HEADER] for b, v in by_bank.items()}
    coverage_ratios = {b: v[NPL_COVERAGE_HEADER] for b, v in by_bank.items()}
    requested = (roc_year, month) if roc_year and month else None
    result = _result("逾期放款總額", year, mon, exact, url, NPL_UNIT, values, requested)
    result["npl_ratios"] = npl_ratios
    result["coverage_ratios"] = coverage_ratios
    return result


def fetch_credit_card_revolving(roc_year=None, month=None, banks=TARGET_BANKS,
                                 cache_dir=_CACHE_DIR, verbose=False):
    """循環信用餘額 (信用卡循環) for `banks`, in 新臺幣千元."""
    links = _list_period_links(CREDIT_CARD_PAGE_URL, _CC_PERIOD_RE, ".zip")
    year, mon, url, exact = resolve_period(links, roc_year, month)
    if verbose:
        print(f"[信用卡循環] using {year}年{mon}月 -> {url}")
    path = _xlsx_from_zip(download_file(url, cache_dir))
    values = _extract_by_bank(path, CC_BANK_HEADER, CC_VALUE_HEADER, banks)
    requested = (roc_year, month) if roc_year and month else None
    return _result("循環信用餘額", year, mon, exact, url, CC_UNIT, values, requested)


# Backwards-compatible alias for this module's original single-dataset entry point.
# ponytail: zero references inside this repo - see DISCLOSURE_PAGE_URL above
# for why it survives Phase 6.
fetch_latest_overdue_loans = fetch_overdue_loans


def quarter_end_month(quarter):
    """Q1->3, Q2->6, Q3->9, Q4->12 - these datasets are monthly, so a
    quarter maps to its END month's publication."""
    return {1: 3, 2: 6, 3: 9, 4: 12}.get(quarter)


def roc_year(western_year):
    return western_year - 1911


def thousands_to_billions(value):
    """新臺幣千元 -> 新臺幣十億元, the unit con-call loan tables use (e.g.
    17,911,768 千元 -> 17.91 十億元). None-safe."""
    return None if value is None else value / 1_000_000


def fetch_for_quarter(western_year, quarter, banks=TARGET_BANKS, cache_dir=_CACHE_DIR, verbose=False):
    """Both datasets for one fiscal quarter, as callfinder.py's con-call
    fallback needs them. Returns {"credit_card": {...}, "overdue": {...}},
    each an individual fetch result dict, with credit_card carrying an
    extra "values_billions" already unit-converted for con-call use.
    Either entry is None if that fetch failed - a network/site problem for
    one dataset shouldn't sink the other, or the whole con-call run."""
    y, m = roc_year(western_year), quarter_end_month(quarter)
    out = {}
    for key, fn in (("credit_card", fetch_credit_card_revolving), ("overdue", fetch_overdue_loans)):
        try:
            out[key] = fn(y, m, banks=banks, cache_dir=cache_dir, verbose=verbose)
        except (RuntimeError, OSError, zipfile.BadZipFile) as e:
            if verbose:
                print(f"[{key}] unavailable: {e}")
            out[key] = None
    if out.get("credit_card"):
        out["credit_card"]["values_billions"] = {
            b: thousands_to_billions(v) for b, v in out["credit_card"]["values"].items()
        }
    return out


def _display_width(s):
    """Terminal column width, counting East-Asian Wide/Fullwidth characters
    as 2 columns - a plain len()-based pad misaligns CJK bank names."""
    return sum(2 if unicodedata.east_asian_width(c) in "WF" else 1 for c in s)


def _print_result(result, extra=None):
    if result is None:
        print("  (unavailable)")
        return
    print(f"\n{result['kind']} - {result['period']} (單位：{result['unit']})")
    if result["note"]:
        print(f"  NOTE: {result['note']}")
    width = max(_display_width(b) for b in result["values"]) + 2
    for bank, value in result["values"].items():
        pad = " " * (width - _display_width(bank))
        shown = f"{value:,.0f}" if isinstance(value, (int, float)) else "N/A（找不到此銀行）"
        if extra and isinstance(value, (int, float)):
            shown += f"   ({extra(value)})"
        print(f"  {bank}{pad}{shown}")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--banks", nargs="*", default=TARGET_BANKS,
                     help="Exact bank legal name(s) to look up (default: the 4 tracked banks).")
    ap.add_argument("--year", type=int, help="Western fiscal year, e.g. 2025 (with --quarter)")
    ap.add_argument("--quarter", type=int, choices=[1, 2, 3, 4], help="Fiscal quarter (with --year)")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    try:
        if args.year and args.quarter:
            both = fetch_for_quarter(args.year, args.quarter, banks=args.banks, verbose=args.verbose)
            cc, npl = both["credit_card"], both["overdue"]
        else:
            cc = fetch_credit_card_revolving(banks=args.banks, verbose=args.verbose)
            npl = fetch_overdue_loans(banks=args.banks, verbose=args.verbose)
    except RuntimeError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    _print_result(cc, extra=lambda v: f"{thousands_to_billions(v):.2f} 十億元")
    _print_result(npl)


if __name__ == "__main__":
    main()
