"""
Extract account values from a folder of converted markdown (.md) financial
statements, using a dedicated account-coding dictionary (an .xlsx, e.g.
"Account Coding.xlsx") instead of keyword/text matching. Since every account
line already has a fixed code (e.g. "19999", "A00010", "3110") tied 1:1 to
its account name, a table row is identified in the source document by its
leading code cell matching the dictionary exactly - no fuzzy/substring
matching is needed.

Given a document and a statement type, this pulls every code+value pair
found in that statement in one pass ("whole statement at once"), rather than
one account per run.

Usage:
    python statements.py <folder> <statement> --coding "Account Coding.xlsx" [--period 2024/12/31] [--export csv] [-v]

<statement> is one of: balance_sheet, income_statement, cash_flow
(the fourth sheet in the coding workbook, the equity statement / 權益變動表,
has a transposed layout - codes as column headers rather than row leaders -
and is not supported by this tool; requesting it raises a clear error).

Requires openpyxl to read the coding dictionary (pip install openpyxl).
"""

import argparse
import csv
import re
import sys
from pathlib import Path
from core.text import despace_cjk, _contains_any, _is_toc_like, page_num, strip_footnote, _strip_footnote_suffix
from core.numbers import parse_numeric, nth_value, format_value, annualize, format_pct, format_maybe_pct
# What THIS module needs from its siblings - the CLI dispatches to all three
# modes, and write_combined_csv emits the ratio columns alongside the
# statement ones. Not a facade: the package __init__ is where the re-exports
# live, so nothing here is imported merely to be re-exposed.
from financialReports.entities import (BANKS, detect_bank, resolve_bank_name,
                                       bank_detection_message)
from financialReports.ratios import (RATIO_COLUMNS, collect_ratio_rows,
                                     print_ratio_rows, write_ratio_csv)
from financialReports.summary import (collect_summary_rows, summary_coverage_warning,
                                      print_summary_rows, write_summary_csv)
from core.industry import INDUSTRY_CODING_FILES, INDUSTRY_CATEGORY_KEYWORDS, detect_industry_category
from core.lookup import find_value_by_label, find_code_value, build_code_index
from core.tables import percent_stride_map, build_raw_lines, _split_dual_column_tables, restrict_section, _is_table_divider, _split_row, _looks_like_code, group_rows_by_code, parse_pipe_tables


def pick_folder():
    """Open a native folder-selection dialog (tkinter, always bundled with
    Python on Windows) and return the chosen path, or None if the user
    cancels. Used as a fallback when <folder> is omitted from the command
    line, so a folder can be picked by clicking instead of typing/pasting
    its full path. Opens in Downloads by default - Windows' native dialog
    otherwise remembers whatever directory some OTHER, unrelated program
    last used it in, which can land somewhere unhelpful (e.g. an old
    project's folder) rather than where these .md conversions actually live."""
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    downloads = Path.home() / "Downloads"
    folder = filedialog.askdirectory(
        title="Select the .md folder",
        initialdir=str(downloads) if downloads.is_dir() else None,
    )
    root.destroy()
    return folder or None

# Windows consoles often default to cp1252, which can't print CJK output.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Statement definitions: coding-workbook sheet name, section markers used to
# isolate that statement's rows within a merged .md document (a single file
# may contain multiple statements back-to-back), and where the code/name
# columns live in that sheet (0-indexed, (original_code_cols, original_name_col,
# revised_code_cols, revised_name_col)).
# ---------------------------------------------------------------------------

STATEMENTS = {
    "balance_sheet": {
        "sheet": "資產負債表",
        "start_markers": ["合併資產負債表", "資產負債表", "Balance Sheet", "BALANCE SHEET"],
        "end_markers": ["合併綜合損益表", "綜合損益表", "Income Statement"],
    },
    "income_statement": {
        "sheet": "綜合損益表",
        "start_markers": ["合併綜合損益表", "綜合損益表", "Income Statement", "Statement of Comprehensive Income"],
        "end_markers": ["合併現金流量表", "現金流量表", "Cash Flow", "CASH FLOW"],
    },
    "cash_flow": {
        "sheet": "現金流量表",
        "start_markers": ["合併現金流量表", "現金流量表", "Cash Flow Statement", "CASH FLOW"],
        "end_markers": ["合併權益變動表", "權益變動表", "Statement of Changes in Equity"],
    },
}

UNSUPPORTED_STATEMENT_MSG = (
    "'equity_statement' (權益變動表) is not supported: its coding sheet is "
    "transposed (codes are column headers, not row leaders), which this "
    "tool's row-per-account matching can't handle."
)

# Industry detection and the coding-workbook table moved to
# core/industry.py - summary mode needs them too. Re-exported above.


def _unmerge_fill(ws):
    """Return a 2D grid of cell values with every merged range's value
    copied into each of its member cells - openpyxl otherwise returns None
    for all but a merge's top-left cell, which would hide most of these
    sheets' header/title text (the industry coding workbooks make heavy
    use of merged header cells spanning several columns)."""
    max_r, max_c = ws.max_row, ws.max_column
    grid = [[ws.cell(row=r, column=c).value for c in range(1, max_c + 1)] for r in range(1, max_r + 1)]
    for mc in ws.merged_cells.ranges:
        val = grid[mc.min_row - 1][mc.min_col - 1]
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                grid[r - 1][c - 1] = val
    return grid


# Header markers distinguishing the "original/pre-revision" code+name block
# from the "revised/post-revision" one (both are present side by side in
# each sheet - a reconciliation between the old and new, IFRS17-era, code
# schemes). Deliberately excludes anything that could ALSO be a real
# account name (e.g. '資產'/'收益' are legitimate level-1 line-item names,
# not just header labels) - an early version of this used those as header
# hints and that caused a real DATA row (code 10000 = '資產') to be misread
# as a header row, pushing the data-start row past it and silently
# dropping that code from the dictionary.
_CODING_ORIG_MARKERS = ["原會計項目及代碼", "修正前會計項目及代碼", "修訂前現金流量表項目及代碼", "修正前"]
_CODING_REV_MARKERS = ["修正後會計項目及代碼", "修訂後現金流量表項目及代碼", "修正後"]
_CODING_HEADER_HINTS = ["代碼", "會計項目", "一級", "二級", "三級", "四級"]
# Columns after these headers are explanatory prose (why a code changed, a
# free-text remark), not code/name data - including them in a block's
# column span lets their long text get picked up as the "name" (via the
# longest-non-code-text heuristic below) instead of the real, much
# shorter, account name.
_CODING_STOP_MARKERS = ["修正說明", "備註"]


def _find_coding_blocks(grid, scan_rows=10):
    """Scan the first `scan_rows` rows for the original-block and revised-
    block column spans (see marker lists above). Returns (orig_span,
    rev_span) as 0-indexed (start, end) inclusive tuples, or None for a
    block absent from this sheet, plus the 0-indexed row where real data
    starts (the row after the last header-hint row found)."""
    n_cols = max(len(r) for r in grid[:scan_rows])
    orig_start = rev_start = stop_col = None
    last_header_row = 0
    for r in range(min(scan_rows, len(grid))):
        row_has_hint = False
        for c in range(n_cols):
            val = grid[r][c]
            if not isinstance(val, str):
                continue
            if any(m in val for m in _CODING_ORIG_MARKERS) and orig_start is None:
                orig_start = c
                row_has_hint = True
            if any(m in val for m in _CODING_REV_MARKERS) and rev_start is None:
                rev_start = c
                row_has_hint = True
            if any(m in val for m in _CODING_STOP_MARKERS) and stop_col is None:
                stop_col = c
                row_has_hint = True
            if val.strip() in _CODING_HEADER_HINTS:
                row_has_hint = True
        if row_has_hint:
            last_header_row = r
    limit = (stop_col - 1) if stop_col is not None else (n_cols - 1)
    orig_end = (rev_start - 1) if (orig_start is not None and rev_start is not None and rev_start > orig_start) else limit
    orig_block = (orig_start, orig_end) if orig_start is not None else None
    rev_block = (rev_start, limit) if rev_start is not None else None
    return orig_block, rev_block, last_header_row + 1


def _extract_coding_block(grid, block, data_start_row):
    """For a (start, end) column span, build {code: name} by scanning each
    data row's cells in that span: the code-shaped cell is the code, and
    the longest non-code text cell is the name (not a fixed column index -
    the name column shifts between rows in at least one observed sheet)."""
    if block is None:
        return {}
    start, end = block
    out = {}
    for row in grid[data_start_row:]:
        cells = row[start:end + 1]
        code = name = None
        for cell in cells:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            if code is None and _looks_like_code(s):
                code = s
            elif not _looks_like_code(s) and (name is None or len(s) > len(name)):
                name = s
        if code and name:
            out[code] = name
    return out


def load_code_dictionary(coding_path, statement):
    """Return {code_str: name} for the given statement, read from an
    industry coding workbook (see INDUSTRY_CODING_FILES). Both the
    original and revised code/name blocks are read; revised entries win
    on conflict since 修正後 is the corrected version. Blank code cells
    are skipped."""
    if openpyxl is None:
        raise RuntimeError("Reading the coding dictionary requires 'openpyxl'. Install with: pip install openpyxl")
    if statement == "equity_statement":
        raise ValueError(UNSUPPORTED_STATEMENT_MSG)
    if statement not in STATEMENTS:
        raise ValueError(f"Unknown statement '{statement}'. Known: {list(STATEMENTS)}")

    spec = STATEMENTS[statement]
    wb = openpyxl.load_workbook(coding_path, data_only=True)
    if spec["sheet"] not in wb.sheetnames:
        raise ValueError(f"Sheet '{spec['sheet']}' not found in {coding_path}. Sheets: {wb.sheetnames}")
    ws = wb[spec["sheet"]]

    grid = _unmerge_fill(ws)
    orig_block, rev_block, data_start_row = _find_coding_blocks(grid)
    orig_dict = _extract_coding_block(grid, orig_block, data_start_row)
    rev_dict = _extract_coding_block(grid, rev_block, data_start_row)
    codes = dict(orig_dict)
    codes.update(rev_dict)
    return codes


def resolve_coding_path(industry=None, folder=None, explicit_path=None):
    """Resolve which coding workbook to load: an explicit path always wins;
    otherwise resolve `industry` (a INDUSTRY_CODING_FILES key) directly, or
    auto-detect it from `folder` if industry is None. Raises ValueError if
    neither an explicit path nor a resolvable industry is available."""
    if explicit_path:
        return explicit_path
    if industry is None and folder is not None:
        industry = detect_industry_category(folder)
    if industry not in INDUSTRY_CODING_FILES:
        raise ValueError(
            f"Couldn't determine an industry coding file (detected: {industry!r}). "
            f"Pass --coding explicitly, or --industry one of {list(INDUSTRY_CODING_FILES)}."
        )
    return INDUSTRY_CODING_FILES[industry]


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def extract_statement(doc_path, statement, code_dict, period=1, verbose=False):
    """Scan doc_path for the given statement's section, match every row
    whose leading cell is a known code, and pull that row's `period`-th
    (1-indexed, most-recent-first) value. Returns a list of dicts:
      {code, name, label_in_doc, value, page_num, confidence}
    (possibly empty if nothing in this file matches).

    If the statement's section marker (e.g. "資產負債表") isn't found in
    this file, the WHOLE file is scanned unrestricted rather than skipped -
    some conversions split one statement per file without repeating its
    title on the data page itself (confirmed against a real 北富銀 filing,
    where the title only appears in the table of contents), so requiring
    the marker would silently produce zero results for an entire folder of
    otherwise-correct data. Matching is by exact code equality regardless,
    so scanning unrestricted doesn't risk false positives the way a looser
    text search would.
    """
    spec = STATEMENTS[statement]
    lines = build_raw_lines(doc_path)
    restricted = restrict_section(lines, spec["start_markers"], spec["end_markers"])
    if restricted is not None:
        lines = restricted

    entries = group_rows_by_code(lines, code_dict)
    results = []
    seen_codes = set()
    for code, page_num, cells, stride in entries:
        if code in seen_codes:
            continue  # keep first match only
        seen_codes.add(code)
        value = nth_value(cells, period, stride)
        results.append({
            "code": code,
            "name": code_dict[code],
            "label_in_doc": cells[1] if len(cells) > 1 else "",
            "value": value,
            "page_num": page_num,
            "confidence": "high" if value is not None else "low",
        })
        if verbose and value is None:
            print(f"  [{code}] matched but no period-{period} value found in row: {cells}")
    return results


def collect_statement_rows(folder, coding, statement, period, verbose=False):
    """Run the account-code extraction for one statement across every .md
    file in `folder`. Returns a list of row dicts (code, name, label_in_doc,
    value, confidence, source_file, page_num). `coding` may be an explicit
    file path, or None to auto-detect the industry category (金控業/金融業/
    保險業) from the filing itself and use its coding workbook."""
    coding = resolve_coding_path(folder=folder, explicit_path=coding)
    code_dict = load_code_dictionary(coding, statement)
    if verbose:
        print(f"Loaded {len(code_dict)} codes for '{statement}' from {coding}")

    rows = []
    for doc_path in sorted(Path(folder).rglob("*.md")):
        results = extract_statement(doc_path, statement, code_dict, period=period, verbose=verbose)
        if results and verbose:
            print(f"[{doc_path.name}] matched {len(results)} code(s)")
        for r in results:
            r["source_file"] = doc_path.name
            rows.append(r)
    return rows
# 'page_num' deliberately excluded: markdown has no page concept, so that
# row field is always None (see build_raw_lines) - the meaningful page
# identifier is the 'page' column derived from source_file via page_num().
STATEMENT_COLUMNS = ["code", "name", "label_in_doc", "value", "confidence"]


def print_statement_rows(statement, rows):
    print(f"\n=== {statement} ===")
    if not rows:
        print("No matching codes found in any file.")
        return
    for r in rows:
        mark = "" if r["confidence"] == "high" else "*"
        print(f"{r['code']}\t{r['name']}\t{format_value(r['value'])}{mark}\t({page_num(r['source_file'])})")


def write_statement_csv(folder, statement, rows):
    out_path = Path(folder) / f"{statement}_export.csv"
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["code", "name", "label_in_doc", "value", "confidence", "page"])
        for r in rows:
            writer.writerow([r["code"], r["name"], r["label_in_doc"], format_value(r["value"]),
                              r["confidence"], page_num(r["source_file"])])
    return out_path


def write_combined_csv(folder, statement_rows_by_name, ratio_rows, used_fallback):
    """Write one combined CSV covering all three statements plus ratios, with
    a leading 'section' column and blank cells for columns that don't apply
    to a given row's section."""
    out_path = Path(folder) / "combined_export.csv"
    header = ["section"] + STATEMENT_COLUMNS + ["page", "period", "entity", "quarter"] + RATIO_COLUMNS
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for statement, rows in statement_rows_by_name.items():
            for r in rows:
                writer.writerow([statement, r["code"], r["name"], r["label_in_doc"], format_value(r["value"]),
                                  r["confidence"], page_num(r["source_file"]), "", "", ""] + [""] * len(RATIO_COLUMNS))
        for r in ratio_rows:
            blank_statement_cols = [""] * len(STATEMENT_COLUMNS)
            if used_fallback:
                # The fallback row carries BOTH annualized figures - see
                # collect_ratio_rows, and both print_ratio_rows and
                # write_ratio_csv emit both. This branch used to emit one
                # value followed by four blanks, which lost ROE entirely and
                # put ROA in the roa_posttax column instead of the annualized
                # one. The manual formula returns plain ratios, so these two
                # format as :.2% rather than through format_pct.
                def fallback_pct(value):
                    return f"{value:.2%}" if value is not None else "N/A"

                ratio_vals = ["",                                        # roa_posttax
                              fallback_pct(r["roa_posttax_annualized"]),
                              "",                                        # roe_posttax
                              fallback_pct(r["roe_posttax_annualized"]),
                              ""]                                        # profit_margin
            else:
                ratio_vals = [format_pct(r[c]) for c in RATIO_COLUMNS]
            writer.writerow(["ratios"] + blank_statement_cols + [page_num(r["source_file"]), r["period"], r["entity"], r["quarter"]] + ratio_vals)
    return out_path


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("folder", nargs="?", default=None,
                     help="Folder containing .md files. If omitted, a folder-picker dialog opens.")
    ap.add_argument("statement", choices=list(STATEMENTS) + ["equity_statement", "ratios", "all", "summary"],
                     help="Which statement to extract, 'ratios' for calculated ROA/ROE, "
                          "or 'all' for balance_sheet + income_statement + cash_flow + ratios together")
    ap.add_argument("--coding", default=None,
                     help="Path to an account-coding dictionary .xlsx, overriding auto-detection. "
                          "By default the industry category (金控業/金融業/保險業) is auto-detected "
                          "from the filing's own entity name and its matching bundled coding file is used.")
    ap.add_argument("--industry", choices=list(INDUSTRY_CODING_FILES),
                     help="Force the industry category instead of auto-detecting it from the filing.")
    ap.add_argument("--period", type=int, default=1,
                     help="Which period to extract, counting left-to-right as listed in the "
                          "document (these filings always list the most recent period first). "
                          "1 = most recent (default), 2 = next most recent, etc.")
    ap.add_argument("--export", choices=["csv"], help="Write results to a CSV file instead of stdout")
    ap.add_argument("--bank",
                     help="Bank name for 'summary' (resolves bank-specific code overrides and "
                          "composite-term formulas). Accepts the short form or full name, e.g. "
                          "either '北富銀' or '台北富邦銀行'. If omitted, auto-detected from the "
                          "filing's own text. Choices: " + ", ".join(BANKS))
    ap.add_argument("-v", "--verbose", action="store_true", help="Print per-file/per-table detail")
    args = ap.parse_args()

    if args.period < 1:
        ap.error("--period must be 1 or greater (1 = the most recent period).")

    if args.statement == "equity_statement":
        # Kept in `choices` so --help still lists it and the user gets THIS
        # message rather than "invalid choice". It used to reach
        # load_code_dictionary and surface as a bare ValueError traceback,
        # despite the docstring promising a clear error.
        ap.error(UNSUPPORTED_STATEMENT_MSG)

    if args.folder is None:
        args.folder = pick_folder()
        if args.folder is None:
            ap.error("No folder selected.")

    if args.statement != "summary":
        # 'summary' never touches a coding dictionary (see collect_summary_rows -
        # it matches raw document codes directly, never the dictionary's name),
        # so industry resolution only matters for the other statement modes.
        try:
            args.coding = resolve_coding_path(industry=args.industry, folder=args.folder, explicit_path=args.coding)
        except ValueError as e:
            ap.error(str(e))
        if args.verbose:
            print(f"Using coding file: {args.coding}")

    if args.statement == "summary":
        if args.bank:
            try:
                bank = resolve_bank_name(args.bank)
            except ValueError as e:
                ap.error(str(e))
        else:
            bank = detect_bank(args.folder)
            if bank is None:
                ap.error(bank_detection_message(args.folder))
            if args.verbose:
                print(f"Auto-detected bank: {bank}")
        try:
            rows = collect_summary_rows(args.folder, bank, period=args.period,
                                         verbose=args.verbose, industry=args.industry)
        except ValueError as e:
            ap.error(str(e))
        # Warned here rather than in print_summary_rows: --export csv returns
        # without ever printing the rows, and that is the path where a failed
        # read is least likely to be noticed.
        warning = summary_coverage_warning(rows)
        if warning:
            print(warning)
        if args.export == "csv":
            out_path = write_summary_csv(args.folder, rows)
            print(f"Wrote {len(rows)} row(s) to {out_path}")
            return
        print_summary_rows(rows)
        return

    if args.statement in ("all", "ratios"):
        if args.bank:
            try:
                bank = resolve_bank_name(args.bank)
            except ValueError as e:
                ap.error(str(e))
        else:
            bank = detect_bank(args.folder)
            if bank is None:
                ap.error(bank_detection_message(args.folder))

    if args.statement == "all":
        statement_rows_by_name = {
            s: collect_statement_rows(args.folder, args.coding, s, args.period, args.verbose)
            for s in STATEMENTS
        }
        ratio_rows, used_fallback = collect_ratio_rows(args.folder, bank, args.coding, args.verbose)

        if args.export == "csv":
            out_path = write_combined_csv(args.folder, statement_rows_by_name, ratio_rows, used_fallback)
            total = sum(len(r) for r in statement_rows_by_name.values()) + len(ratio_rows)
            print(f"Wrote {total} row(s) to {out_path}")
            return

        for statement, rows in statement_rows_by_name.items():
            print_statement_rows(statement, rows)
        print_ratio_rows(ratio_rows, used_fallback)
        return

    if args.statement == "ratios":
        rows, used_fallback = collect_ratio_rows(args.folder, bank, args.coding, args.verbose)
        if args.export == "csv":
            out_path = write_ratio_csv(args.folder, rows, used_fallback)
            print(f"Wrote {len(rows)} row(s) to {out_path}")
            return
        print_ratio_rows(rows, used_fallback)
        return

    rows = collect_statement_rows(args.folder, args.coding, args.statement, args.period, args.verbose)
    if args.export == "csv":
        out_path = write_statement_csv(args.folder, args.statement, rows)
        print(f"Wrote {len(rows)} row(s) to {out_path}")
        return
    print_statement_rows(args.statement, rows)


if __name__ == "__main__":
    main()
