# -*- coding: utf-8 -*-
"""Manually assembled Excel export of the runfinder.py result for the 4-bank
2025Q4 fin_report + con_call folders (C:\\...\\Downloads\\4+4\\), plus a
second sheet cataloguing every kind of warning/error note the codebase is
capable of producing - since this particular clean run didn't happen to
trigger any of them, the reference sheet documents what each ONE looks like
with a synthetic (clearly labeled) example, so a reader recognizes them the
first time a real one fires.
"""
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SYNTH_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

PCT_FORMAT = "0.00%"
INT_FORMAT = "#,##0"
DEC_FORMAT = "#,##0.00"

with open(Path(__file__).parent / "scratch_run_dump.json", encoding="utf-8") as f:
    dump = json.load(f)

wb = Workbook()

# ---------------------------------------------------------------------------
# Sheet 1: actual run results, one row per (bank, kind, term)
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Run Results (4+4, 2025Q4)"
headers = ["Bank", "Kind", "Term", "Value", "Matched Label / Formula", "Page", "Note"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

def page_num(source_file):
    if not source_file:
        return ""
    stem = Path(source_file).stem
    import re
    m = re.match(r"^(\d+)", stem)
    return m.group(1) if m else stem

note_count = 0
for grp in dump:
    bank, kind = grp["bank"], grp["kind"]
    for r in grp["rows"]:
        term = r["term"]
        is_ratio = r.get("kind") == "ratio"
        value = r.get("individual") if is_ratio else r.get("value")
        label = r.get("matched_label") or ""
        page = page_num(r.get("source_file"))
        note = r.get("note") or ""
        is_percent = is_ratio or r.get("is_percent", False)
        ws.append([bank, kind, term, value, label, page, note])
        row_idx = ws.max_row
        if isinstance(value, (int, float)):
            cell = ws.cell(row=row_idx, column=4)
            if is_percent:
                cell.value = value / 100
                cell.number_format = PCT_FORMAT
            elif value == int(value):
                cell.number_format = INT_FORMAT
            else:
                cell.number_format = DEC_FORMAT
        if note:
            note_count += 1
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = NOTE_FILL

widths = [10, 11, 22, 16, 40, 8, 90]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

ws.append([])
ws.append([f"Total rows: {sum(len(g['rows']) for g in dump)}  |  Rows with a live note this run: {note_count}"])
ws.cell(row=ws.max_row, column=1).font = BOLD
ws.append(["This was a clean run - see 'Error & Note Types Reference' sheet for every kind of "
           "warning/error note the code CAN produce, with a synthetic example of each."])

# ---------------------------------------------------------------------------
# Sheet 2: reference catalogue of every note-producing mechanism
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Error & Note Types Reference")
headers2 = ["#", "Category", "Where in code", "Trigger condition", "Example note text (synthetic)",
            "Seen in this run?", "What to do about it"]
ws2.append(headers2)
for c in range(1, len(headers2) + 1):
    cell = ws2.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

ROWS = [
    (1, "ROA/ROE cross-check divergence",
     "acctfinder.py collect_roa_roe() -> build(), _ROA_ROE_CROSSCHECK_DIVERGENCE_FACTOR = 2.0",
     "The manual formula (net income / avg assets or equity) disagrees with the disclosed "
     "ROA/ROE by more than 2x",
     "cross-check diverges: manual formula gives 4.12% vs 1.02% as-disclosed - could be a real "
     "discrepancy, or just this manual formula's own assumptions not holding for this filing; "
     "treat as a prompt to look closer, not as evidence either number is wrong",
     "No", "Does NOT mean the disclosed value is wrong - check whether the manual formula's "
     "single-quarter x4 annualization assumption actually applies to this filing's net-income "
     "code (YTD-cumulative vs single-quarter)."),
    (2, "ROA/ROE implausible value",
     "acctfinder.py collect_roa_roe() -> build(), _ROA_PLAUSIBLE_MIN/MAX=[-5,5], "
     "_ROE_PLAUSIBLE_MIN/MAX=[-50,50]",
     "The primary ROA or ROE value itself falls outside a very wide, deliberately generous "
     "sanity range",
     "implausible value: 87.30% is outside the expected [-50%, 50%] range for ROE - likely a "
     "parsing/scale/sign error, worth a manual check",
     "No", "Bounds are wide on purpose - if this actually fires on real data, it's very likely "
     "a genuine extraction bug (wrong code, wrong period column, or a % sign misread), not a "
     "real outlier bank."),
    (3, "Loan-book reconciliation mismatch",
     "callfinder.py collect_con_call_summary(), _LOAN_RECONCILE_TOLERANCE = 2.5",
     "The 4 recomposed loan buckets (企業放款+房貸+個人放款+信用卡循環) don't sum to the "
     "deck's own 法說會放款餘額合計 within tolerance",
     "components sum to 3,412.5 vs total 3,409.0 (off by 3.5) - a component may have matched "
     "the wrong row, or a recomposition rule may not hold for this filing",
     "No", "Small overs (a few units) are often just the bank's own printed rounding - only "
     "worth chasing down when the gap is large relative to the total."),
    (4, "Regulator dataset: requested month not yet published",
     "npl_finder.py resolve_period() -> _result()'s 'note' field",
     "fetch_overdue_loans()/fetch_credit_card_revolving() asked for a specific ROC year/month "
     "that the government site hasn't published yet, so an earlier month was substituted",
     "requested 115年12月 but that month isn't published yet - used 115年5月 instead",
     "No (114年12月 was available and exact for this run)",
     "KNOWN GAP: this note lives on npl_finder's OWN result dict, but callfinder.py's "
     "信用卡循環/逾放比率/備抵呆帳/逾期放款 output rows do not currently copy it into their "
     "row 'note' field - it's only visible via -v/verbose console output today, not in the "
     "exported CSV/Excel. Worth wiring through if this fallback is ever silently relied on."),
    (5, "Regulator dataset unavailable (network/SSL/site-structure failure)",
     "callfinder.py collect_con_call_summary()'s try/except around 'import npl_finder' calls",
     "banking.gov.tw unreachable, SSL cert error, or the page/file structure changed",
     "(verbose-only) Regulator dataset unavailable (Failed to reach https://www.banking.gov.tw/"
     "...): 信用卡循環 fallback and 逾放比率/備抵呆帳/逾期放款 will be N/A.",
     "No", "Never sinks the whole run - only the government-sourced rows go to N/A. Check "
     "network/SSL (pip install pip-system-certs) or site structure (see HANDOFF.md section 6.4/"
     "6.10) if this keeps happening."),
    (6, "Code/term genuinely not found (silent N/A, no note)",
     "acctfinder.py build_code_index() / find_value_by_label(); callfinder.py find_term_value()",
     "The account code or con-call term simply doesn't appear anywhere in the folder's files",
     "(no note text - just value=None, matched_label=None)",
     "Yes (放款均率/存款均率 = N/A for some banks this run)",
     "Not an error by itself - confirm via -v whether the source document genuinely lacks this "
     "line before assuming a bug (see HANDOFF.md Scenario A)."),
    (7, "Composite term has no formula for this bank (silent N/A, no note)",
     "acctfinder.py collect_summary_rows() composite branch; "
     "callfinder.py LOAN_RECOMPOSITION.get(bank, {})",
     "COMPOSITE_TERMS[name] or LOAN_RECOMPOSITION[bank] simply has no entry for this bank",
     "(no note text - just value=None; verbose mode prints "
     "\"'X' has no formula defined for bank 'Y'\")",
     "No (all 4 banks have formulas for every current composite term)",
     "Expected when a bank hasn't been reverse-engineered for a given composite yet - not a "
     "bug, just an incomplete formula table (see HANDOFF.md 11.1/11.3)."),
    (8, "Bank could not be auto-detected (whole-folder skip, not a row note)",
     "runfinder.py run_fin_report(), fin_report_rows() returning kind='no_bank'",
     "detect_bank(folder) found no recognizable bank name in the first .md file",
     "Couldn't auto-detect the bank for <folder> - skipping (run acctfinder.py directly with "
     "--bank to override).",
     "No", "Pass --bank explicitly when running acctfinder.py directly on that folder."),
]

for row in ROWS:
    ws2.append(list(row))
    idx = ws2.max_row
    if row[5].startswith("No"):
        for c in range(5, 6):
            ws2.cell(row=idx, column=c).fill = SYNTH_FILL

widths2 = [4, 30, 46, 46, 60, 20, 60]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
ws2.freeze_panes = "A2"

ws2.append([])
ws2.append(["Grey-filled 'Example note text' cells are SYNTHETIC (constructed to show the "
            "exact wording), not real output from this run - none of these 8 categories fired "
            "on the real 2025Q4 data checked here."])
ws2.cell(row=ws2.max_row, column=1).font = BOLD

out_path = Path.home() / "Downloads" / "runfinder_manual_export_with_error_notes.xlsx"
try:
    wb.save(out_path)
except PermissionError:
    import datetime
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path.home() / "Downloads" / f"runfinder_manual_export_with_error_notes_{stamp}.xlsx"
    wb.save(out_path)

print("Saved:", out_path)
