# -*- coding: utf-8 -*-
"""Fully FICTIONAL demo export - a made-up bank ('範例商業銀行'), a made-up
period (1Q26), and made-up numbers throughout. Nothing here is a real filing
or real deck. The numbers are deliberately chosen so several of them
naturally trigger the code's own note-generating logic (ROA cross-check
divergence, ROE implausibility, loan-book reconciliation mismatch, regulator
month-fallback) - i.e. the notes below are exactly what acctfinder.py/
callfinder.py would print for these inputs, not text pasted in afterward.
"""
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PCT_FORMAT = "0.00%"
INT_FORMAT = "#,##0"
DEC_FORMAT = "#,##0.00"
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BOLD = Font(bold=True)

# ---------------------------------------------------------------------------
# FICTIONAL fin_report rows for 範例商業銀行, 1Q26 (made-up individual filing)
# ---------------------------------------------------------------------------
# 淨收益 components (61,200,150 + 41,780,320 + 15,900,410 + 13,570,000) sum
# exactly to 淨收益 (132,450,880) - fabricated to look internally consistent.
# 營業費用 components (37,200,000 + 5,100,000 + 26,600,000) sum to 68,900,000.
# 稅前淨利 = 132,450,880 - 68,900,000 - 8,750,000(呆帳) = 54,800,880.
# 稅後淨利 = 54,800,880 - 8,220,000(稅) = 46,580,880.
FIN_ROWS = [
    ("總資產", 6120884331, "資產總計", "007", False, ""),
    ("淨收益", 132450880, "淨收益", "008", False, ""),
    ("利息淨收益", 61200150, "利息淨收益", "008", False, ""),
    ("手續費淨收益", 41780320, "手續費淨收益", "008", False, ""),
    ("評價及已實現", 15900410, "評價及已實現", "008", False, ""),
    ("其他非利息收益", 13570000, "其他非利息收益", "008", False, ""),
    ("營業費用", 68900000, "營業費用合計", "008", False, ""),
    ("員工福利費用", 37200000, "員工福利費用", "008", False, ""),
    ("折舊及攤銷費用", 5100000, "折舊及攤銷費用", "008", False, ""),
    ("其他費用", 26600000, "其他業務及管理費用", "008", False, ""),
    ("呆帳提存(收回)", 8750000, "呆帳費用、承諾及保證責任準備提存", "008", False, ""),
    ("稅前淨利", 54800880, "繼續營業單位稅前淨利", "008", False, ""),
    ("所得稅費用", 8220000, "減：所得稅費用", "008", False, ""),
    ("稅後淨利", 46580880, "本期淨利", "008", False, ""),
    # 活存比: genuinely not disclosed by this fictional filing either -
    # value=None, no note, matches real N/A behavior (category 6 in the
    # error-notes reference, not a warning by itself).
    ("活存比", None, None, "", True, ""),
    # CIR: computed directly, no crosscheck (per current design) -
    # abs(68,900,000) / 132,450,880 * 100 = 52.02%.
    ("CIR", 52.02, "abs(營業費用) / 淨收益", "", True, ""),
    # ROA: disclosed (as-filed) value is 0.95%, but the manual formula
    # (net income / avg assets, x4 annualized) comes out to 3.98% - more
    # than 2x apart -> triggers the cross-check-diverges note (acctfinder.py
    # collect_roa_roe(), _ROA_ROE_CROSSCHECK_DIVERGENCE_FACTOR=2.0). The
    # PRIMARY value shown is still the disclosed 0.95% - the crosscheck
    # never overrides it.
    ("ROA", 0.95, "ROA(年) 稅後 @ 115年3月31日", "131",
     True, "cross-check diverges: manual formula gives 3.98% vs 0.95% as-disclosed - could be "
           "a real discrepancy, or just this manual formula's own assumptions not holding for "
           "this filing; treat as a prompt to look closer, not as evidence either number is wrong"),
    # ROE: the DISCLOSED value itself is implausible (outside [-50%,+50%]) -
    # triggers the plausibility-bound note directly on the primary value
    # (acctfinder.py _ROE_PLAUSIBLE_MIN/MAX), a different code path than ROA's.
    ("ROE", 68.40, "ROE(年) 稅後 @ 115年3月31日", "131",
     True, "implausible value: 68.40% is outside the expected [-50%, 50%] range for ROE - "
           "likely a parsing/scale/sign error, worth a manual check"),
]

# ---------------------------------------------------------------------------
# FICTIONAL con_call rows for the same fictional bank, same fictional period
# ---------------------------------------------------------------------------
# Loan components (1,540 + 1,250 + 310.5 + 16.2 = 3,116.7) vs the deck's own
# printed total (3,120.0) - off by 3.3, past _LOAN_RECONCILE_TOLERANCE=2.5 -
# triggers the reconciliation-mismatch note on the total row.
# 信用卡循環/逾放比率/備抵呆帳/逾期放款 all fictionally requested 115年12月
# (period-end of this fictional 1Q26... no, matches Q4 request pattern) but
# that month "isn't published yet" on the (fictional) government site, so a
# fallback to 115年9月 was used - the note that SHOULD show here demonstrates
# the currently-unwired gap flagged in the previous export (real code today
# only prints this via -v, doesn't put it in the row's own note field yet).
NPL_FALLBACK_NOTE = ("requested 115年12月 but that month isn't published yet - used 115年9月 "
                      "instead [demonstrating a currently-unwired note - see HANDOFF.md/prior "
                      "export's Error & Note Types Reference, category 4]")

CALL_ROWS = [
    ("NIM", 1.58, "淨利息收益率", "048", True, ""),
    ("放款均率", 3.21, "放款利率", "048", True, ""),
    ("存款均率", 1.35, "存款利率", "048", True, ""),
    ("存放利差", 1.86, "整體利差", "017", True, ""),
    ("企業放款", 1540.0, "重組：台幣法人放款 + 外幣放款 − 海外子行", "013", False, ""),
    ("房貸", 1250.0, "房屋貸款", "013", False, ""),
    ("個人放款", 310.5, "重組：信用貸款與其他 − 信用卡循環", "", False, ""),
    ("信用卡循環", 16.2, "循環信用餘額（金管會115年9月）", "", False, NPL_FALLBACK_NOTE),
    ("法說會放款餘額合計", 3120.0, "重組：總放款 − 海外子行", "013", False,
     "components sum to 3,116.7 vs total 3,120.0 (off by -3.3) - a component may have matched "
     "the wrong row, or a recomposition rule may not hold for this filing"),
    ("法說會外幣放款", 610.0, "重組：海外分行 + OBU+DBU", "013", False, ""),
    ("逾放比率", 0.19, "金管會公布（115年9月）", "", True, NPL_FALLBACK_NOTE),
    ("備抵呆帳/逾期放款", 812.40, "金管會公布（115年9月）", "", True, NPL_FALLBACK_NOTE),
]

# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "FICTIONAL demo (範例商業銀行, 1Q26)"

ws.append(["*** ALL DATA ON THIS SHEET IS FICTIONAL - fabricated for demonstration only ***"])
ws.cell(row=1, column=1).font = Font(bold=True, color="C00000", size=12)
ws.append([])

def write_section(title, rows):
    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    header_row = ws.max_row + 1
    ws.append(["Term", "Value", "Matched Label / Formula", "Page", "Note"])
    for c in range(1, 6):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for term, value, label, page, is_percent, note in rows:
        ws.append([term, value, label or "", page, note])
        idx = ws.max_row
        if isinstance(value, (int, float)):
            cell = ws.cell(row=idx, column=2)
            if is_percent:
                cell.value = value / 100
                cell.number_format = PCT_FORMAT
            elif value == int(value):
                cell.number_format = INT_FORMAT
            else:
                cell.number_format = DEC_FORMAT
        if note:
            for c in range(1, 6):
                ws.cell(row=idx, column=c).fill = NOTE_FILL
    ws.append([])

write_section("fin_report (虛構財報, SUMMARY_LAYOUT order)", FIN_ROWS)
write_section("con_call (虛構法說會)", CALL_ROWS)

widths = [22, 16, 46, 8, 95]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows():
    for cell in row:
        if cell.column == 5:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.append(["Notes triggered in this fictional dataset (all by the numbers chosen above, not "
           "hand-written afterward):"])
ws.cell(row=ws.max_row, column=1).font = BOLD
ws.append(["1. ROA: disclosed 0.95% vs manual-formula 3.98% -> cross-check divergence note "
           "(both numbers real outputs of the actual formulas, given the fictional inputs)."])
ws.append(["2. ROE: disclosed value itself (68.40%) exceeds the [-50%,+50%] plausibility bound "
           "-> implausible-value note."])
ws.append(["3. 法說會放款餘額合計: recomposed components (1,540+1,250+310.5+16.2=3,116.7) vs "
           "deck total (3,120.0), off by 3.3 > tolerance 2.5 -> reconciliation-mismatch note."])
ws.append(["4. 信用卡循環/逾放比率/備抵呆帳/逾期放款: fictionally requested 115年12月, not yet "
           "published, fell back to 115年9月 -> fallback note (shown here to illustrate the "
           "wording; the REAL code today only prints this via -v, doesn't attach it to the row "
           "yet - see prior export's reference sheet, category 4)."])
ws.append(["5. 活存比: genuinely not disclosed even in this fictional filing -> blank/N/A, no "
           "note (this is normal, not a warning)."])

out_path = Path.home() / "Downloads" / "fictional_bank_demo_export.xlsx"
try:
    wb.save(out_path)
except PermissionError:
    import datetime
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path.home() / "Downloads" / f"fictional_bank_demo_export_{stamp}.xlsx"
    wb.save(out_path)

print("Saved:", out_path)
